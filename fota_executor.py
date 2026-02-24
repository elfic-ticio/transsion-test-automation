"""
FOTA Test Cases Executor para Claro Colombia
50 test cases basados en TestCase_FOTA_Claro_V1.4
"""
import threading
import time
import json
import os
import logging
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional, List, Dict, Tuple

logger = logging.getLogger(__name__)


@dataclass
class FOTATestCase:
    id: int
    title: str
    description: str
    expected: str
    automation: str  # 'auto', 'semi', 'manual'
    auto_func: str = ""  # nombre del método _test_*
    default_result: str = "pending"  # 'pending' o 'NA'

    def to_dict(self):
        return asdict(self)


# ==================== DEFINICIÓN DE 50 TEST CASES ====================

FOTA_TEST_CASES = [
    FOTATestCase(1, "SW Version",
        "1. Factory reset\n2. Comparar SW version en UI con la del test plan y código interno",
        "SW version debe coincidir. Número de versión debe ser +1 de la anterior",
        "auto", "_test_sw_version"),

    FOTATestCase(2, "Network name",
        "1. Verificar que el nombre de red se muestra correctamente para el operador",
        "Nombre de red correcto (SPN)",
        "auto", "_test_network_name"),

    FOTATestCase(4, "Voice mail",
        "1. Verificar buzón de voz con tecla 1\n2. Marcar número directo\n3. Verificar ícono de voicemail",
        "DUT debe conectar al buzón de voz y mostrar número correctamente",
        "semi", "_test_voicemail"),

    FOTATestCase(5, "Emergency call (sin SIM)",
        "Sin SIM insertada: Realizar llamada de emergencia (911, 112, 123)",
        "Números de emergencia accesibles y conectan al servicio correcto",
        "manual", default_result="pending"),

    FOTATestCase(6, "Emergency call (con SIM)",
        "Con SIM insertada: Realizar llamada de emergencia (911, 112, 123)",
        "Números de emergencia accesibles y conectan al servicio correcto",
        "manual"),

    FOTATestCase(7, "MMS",
        "Enviar/Recibir MMS con imagen, sonido, video y texto",
        "MMS enviados/recibidos correctamente sin contenido faltante",
        "semi", "_test_mms"),

    FOTATestCase(8, "SMS",
        "1. Crear SMS con caracteres especiales\n2. Verificar contador de caracteres\n3. Enviar/recibir SMS",
        "Caracteres correctos, contador correcto, SMS enviado/recibido",
        "auto", "_test_sms"),

    FOTATestCase(9, "Browser",
        "1. Conectar a homepage\n2. Verificar bookmarks por defecto",
        "Browser conecta, navega, descarga y almacena contenido",
        "semi", "_test_browser"),

    FOTATestCase(10, "SIM (STK)",
        "1. Verificar nombre del ícono STK\n2. Verificar acceso a opciones STK",
        "DUT muestra el nombre STK correctamente y accede a las opciones",
        "semi", "_test_stk"),

    FOTATestCase(11, "Radio FM",
        "1. Acceder a Radio FM\n2. Verificar escaneo automático de estaciones",
        "Radio FM muestra estaciones disponibles y reproduce",
        "manual"),

    FOTATestCase(12, "Idle",
        "Verificar: Barra indicadora, RSSI, Wallpaper, Hora/Fecha, Batería, Backlight, Lockscreen",
        "Todos los elementos funcionan correctamente",
        "semi", "_test_idle"),

    FOTATestCase(13, "Menu Tree",
        "1. Entrar a cada menú\n2. Volver a idle después de cada uno",
        "Sin animaciones rotas, sin traducciones incorrectas",
        "semi", "_test_menu_tree"),

    FOTATestCase(14, "Power ON/OFF",
        "1. Apagar/encender el DUT\n2. Verificar animaciones y textos",
        "Animaciones correctas, sin errores de calidad",
        "auto", "_test_power_onoff"),

    FOTATestCase(15, "VoLTE icon",
        "1. Encender DUT\n2. Verificar ícono VoLTE conectado",
        "Ícono VoLTE visible en la barra de estado",
        "auto", "_test_volte_icon"),

    FOTATestCase(16, "VoLTE airplane",
        "1. Verificar ícono VoLTE\n2. Activar modo avión\n3. Desactivar modo avión",
        "VoLTE desaparece con avión, reaparece al desactivar",
        "auto", "_test_volte_airplane"),

    FOTATestCase(17, "VoLTE call",
        "Llamada MO VoLTE a: VoLTE phone, Línea fija, Non-VoLTE\n"
        "Verificar: ringback, conexión, calidad de voz, speaker, etiqueta 'Llamada VoLTE'",
        "Llamada establecida, calidad normal, sin eco/ruido, DUT permanece en LTE",
        "semi", "_test_volte_call"),

    FOTATestCase(18, "VoLTE Emergency call",
        "1. Con VoLTE activo\n2. Llamada de emergencia (911, 112)",
        "Llamada de emergencia exitosa (VoLTE o CSFB a 3G)",
        "manual"),

    FOTATestCase(19, "VoLTE conference",
        "1. Llamada A→B\n2. Agregar llamada A→C\n3. Merge\n4. Split",
        "Conferencia establecida correctamente, split funciona",
        "semi", "_test_volte_conference"),

    FOTATestCase(20, "VoLTE conference (B ends)",
        "1. Llamada A→B\n2. Agregar A→C\n3. Merge\n4. B cuelga",
        "Conferencia OK. Al colgar B, A y C quedan en llamada normal",
        "semi", "_test_volte_conference_b_ends"),

    FOTATestCase(21, "ViLTE",
        "Videollamada ViLTE: verificar calidad de audio y video",
        "Videollamada conectada, calidad normal de audio y video",
        "manual", default_result="NA"),

    FOTATestCase(22, "VoLTE SMS/MMS/Email",
        "Enviar y recibir SMS/MMS/Email durante VoLTE activo",
        "SMS/MMS/Email enviados y recibidos correctamente",
        "auto", "_test_volte_sms"),

    FOTATestCase(23, "VoWiFi registration",
        "1. Apagar y encender dispositivo\n2. Verificar registro VoWiFi",
        "DUT registra en VoWiFi (verificar ícono según spec del operador)",
        "auto", "_test_vowifi_registration"),

    FOTATestCase(24, "VoWiFi airplane",
        "1. Activar modo avión\n2. Habilitar WiFi\n3. Observar registro IMS",
        "Registro IMS debe realizarse",
        "auto", "_test_vowifi_airplane"),

    FOTATestCase(25, "VoWiFi call",
        "1. Llamada MO por VoWiFi\n2. Aceptar en Phone B\n3. Verificar calidad de voz",
        "Calidad normal, sin eco, sin ruido, sin delay",
        "semi", "_test_vowifi_call"),

    FOTATestCase(26, "VoWiFi Emergency call",
        "Llamada de emergencia por VoWiFi (911, 112)",
        "Llamada de emergencia exitosa",
        "manual"),

    FOTATestCase(27, "VoWiFi SMS/MMS/Email",
        "Enviar y recibir SMS/MMS/Email por VoWiFi",
        "SMS/MMS/Email enviados y recibidos correctamente",
        "manual"),

    FOTATestCase(28, "Call types",
        "MO, MO Video, MO Internacional, MT, MT Video, Missed, Missed Video, Call history",
        "Menú correcto, categorías e íconos correctos, hora/fecha correcta",
        "semi", "_test_call_types"),

    FOTATestCase(29, "Multiparty MO",
        "1. MO call\n2. Segunda MO call durante primera\n3. Tercera MO call\n4. Verificar estado",
        "Tercera llamada funciona, swap y transferencia OK",
        "semi", "_test_multiparty_mo"),

    FOTATestCase(30, "Multiparty MT",
        "1. MO call\n2. Recibir MT durante llamada\n3. Join + recibir nueva MT\n4. Verificar estado",
        "Tercera llamada funciona, swap y transferencia OK",
        "semi", "_test_multiparty_mt"),

    FOTATestCase(31, "Email",
        "1. Configurar cuenta (Exchange, Google, Yahoo, Hotmail)\n2. Enviar multimedia\n3. Sincronizar",
        "Emails enviados, recibidos y mostrados correctamente",
        "manual"),

    FOTATestCase(32, "WiFi",
        "1. Buscar redes WLAN\n2. Conectar a una red\n3. Desconectar y reconectar",
        "Handset encuentra y conecta a redes WLAN correctamente",
        "auto", "_test_wifi"),

    FOTATestCase(33, "WiFi Streaming",
        "1. Abrir YouTube y reproducir video 1 min\n2. Abrir browser a www.wwe.com",
        "Video y página web cargan correctamente",
        "semi", "_test_wifi_streaming"),

    FOTATestCase(34, "Bluetooth search",
        "1. Habilitar BT\n2. Buscar dispositivos\n3. Comparar con referencia",
        "Resultados de búsqueda coinciden con teléfono de referencia",
        "auto", "_test_bluetooth_search"),

    FOTATestCase(35, "Bluetooth transfer",
        "1. Habilitar BT\n2. Enviar/recibir imagen, sonido, contacto, MP3\n3. Verificar transferencia",
        "Archivos enviados y recibidos correctamente vía BT",
        "manual"),

    FOTATestCase(36, "Alarm",
        "1. Configurar alarma\n2. Probar todas las repeticiones (Once, Daily, Mon-Fri, etc.)",
        "Alarma suena a tiempo, snooze funciona",
        "auto", "_test_alarm"),

    FOTATestCase(37, "Calculator",
        "1. Verificar todas las operaciones disponibles",
        "Resultados correctos, idioma correcto, división por cero manejada",
        "auto", "_test_calculator"),

    FOTATestCase(38, "Camera",
        "1. Abrir cámara\n2. Cambiar opciones\n3. Tomar foto",
        "Opciones de cámara funcionan correctamente, idioma correcto",
        "auto", "_test_camera"),

    FOTATestCase(39, "MP3",
        "1. Reproducir MP3\n2. Cambiar opciones\n3. Verificar reproducción",
        "Opciones de MP3 funcionan correctamente",
        "manual"),

    FOTATestCase(40, "Ear-Mic",
        "1. Conectar auriculares\n2. Llamada\n3. Cambiar volumen\n4. Reproducir música",
        "Volumen cambia correctamente",
        "manual"),

    FOTATestCase(41, "Save contact",
        "1. Crear contacto en: cuenta Google, almacenamiento, SIM",
        "Contacto creado, se puede eliminar, copiar y compartir",
        "auto", "_test_save_contact"),

    FOTATestCase(42, "External memory",
        "1. Insertar memoria externa\n2. Mover/copiar/borrar archivos",
        "Memoria funciona, carpetas por defecto creadas, archivos no corruptos",
        "manual"),

    FOTATestCase(43, "Customization",
        "1. Verificar versión de apps preinstaladas (AMX/Claro)",
        "Apps con la versión del AVISO implementada en la TA",
        "auto", "_test_customization"),

    FOTATestCase(44, "SYSDLL",
        "1. Verificar app SYSDLL preinstalada con versión correcta vía ADB",
        "SYSDLL con versión del AVISO implementada en la TA",
        "auto", "_test_sysdll"),

    FOTATestCase(45, "FOTA preservation",
        "Antes de FOTA: personalizar (wallpaper, íconos, ringtone, instalar top 10 apps)\n"
        "Después de FOTA: verificar que personalización se mantiene",
        "DUT mantiene toda la personalización del usuario",
        "semi", "_test_fota_preservation"),

    FOTATestCase(46, "FOTA popup",
        "1. Encender DUT\n2. Verificar idioma local\n3. Verificar popup FOTA",
        "Popup FOTA se muestra correctamente",
        "semi", "_test_fota_popup"),

    FOTATestCase(47, "FOTA SW version",
        "1. Verificar SW Version en UI\n2. Verificar SW Version en menú de ingeniería",
        "SW version coincide",
        "auto", "_test_fota_sw_version"),

    FOTATestCase(48, "FOTA wordings",
        "1. Verificar todos los textos del proceso FOTA",
        "Textos en español y no cortados",
        "semi", "_test_fota_wordings"),

    FOTATestCase(49, "FOTA low battery",
        "1. Intentar FOTA con batería baja\n2. Verificar textos",
        "No permite FOTA. Debe cargar más del 30%",
        "manual"),

    FOTATestCase(50, "FOTA Automatically",
        "1. Intentar FOTA con red de datos",
        "DUT no permite descarga por datos. Solo por WiFi",
        "semi", "_test_fota_auto"),

    FOTATestCase(51, "Network 2G",
        "Verificar conectividad 2G",
        "Conectividad 2G funcional",
        "manual", default_result="NA"),
]


class FOTAExecutor:
    """Ejecutor de test cases FOTA para Claro"""

    def __init__(self, adb_manager):
        self.adb = adb_manager
        self.test_cases = {tc.id: tc for tc in FOTA_TEST_CASES}
        self.results: Dict[int, Dict] = {}
        self.logs: List[str] = []
        self.is_running = False
        self.progress = {
            'current_test': 0,
            'total_tests': 0,
            'phase': 'idle',
            'status': 'idle'
        }
        self._stop_flag = threading.Event()
        self._lock = threading.Lock()
        self._thread: Optional[threading.Thread] = None

        # Inicializar resultados
        for tc in FOTA_TEST_CASES:
            self.results[tc.id] = {
                'result': tc.default_result,
                'remark': '',
                'screenshot': None,
                'timestamp': None
            }

    def _log(self, message: str, level: str = "INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        entry = f"[{ts}] [{level}] {message}"
        with self._lock:
            self.logs.append(entry)
            if len(self.logs) > 200:
                self.logs = self.logs[-200:]
        logger.info(message) if level == "INFO" else logger.error(message)

    # ==================== API PÚBLICA ====================

    def get_test_cases(self) -> List[Dict]:
        """Retorna todos los test cases con sus resultados"""
        result = []
        for tc in FOTA_TEST_CASES:
            d = tc.to_dict()
            d['result'] = self.results[tc.id]['result']
            d['remark'] = self.results[tc.id]['remark']
            d['screenshot'] = self.results[tc.id]['screenshot']
            result.append(d)
        return result

    def get_state(self) -> Dict:
        with self._lock:
            counts = {'Pass': 0, 'Fail': 0, 'NA': 0, 'pending': 0}
            for r in self.results.values():
                res = r['result']
                if res in counts:
                    counts[res] += 1
                else:
                    counts['pending'] += 1
            return {
                'is_running': self.is_running,
                'progress': self.progress.copy(),
                'counts': counts,
                'logs': self.logs[-50:],
                'results': {k: v.copy() for k, v in self.results.items()}
            }

    def run_single(self, serial: str, test_id: int, dut2_serial: str = None, dut2_phone: str = None) -> Dict:
        """Ejecuta un solo test case (síncrono)"""
        tc = self.test_cases.get(test_id)
        if not tc:
            return {'result': 'Fail', 'remark': 'Test case no encontrado'}

        if tc.automation == 'manual' and tc.default_result != 'NA':
            return {'result': 'pending', 'remark': 'Test manual - requiere intervención del tester'}

        if not tc.auto_func:
            return {'result': 'pending', 'remark': 'Sin función de automatización'}

        func = getattr(self, tc.auto_func, None)
        if not func:
            return {'result': 'Fail', 'remark': f'Función {tc.auto_func} no implementada'}

        self._log(f"Ejecutando test #{tc.id}: {tc.title}")
        try:
            result = func(serial, dut2_serial, dut2_phone)
            self.results[tc.id].update({
                'result': result.get('result', 'Fail'),
                'remark': result.get('remark', ''),
                'screenshot': result.get('screenshot'),
                'timestamp': datetime.now().isoformat()
            })
            self._log(f"Test #{tc.id}: {result.get('result', 'Fail')} - {result.get('remark', '')}")
            return result
        except Exception as e:
            err = f"Error ejecutando test #{tc.id}: {str(e)}"
            self._log(err, "ERROR")
            self.results[tc.id].update({
                'result': 'Fail',
                'remark': str(e),
                'timestamp': datetime.now().isoformat()
            })
            return {'result': 'Fail', 'remark': str(e)}

    def run_all_auto(self, serial: str, dut2_serial: str = None, dut2_phone: str = None) -> bool:
        """Ejecuta todos los tests automáticos en un thread"""
        if self.is_running:
            return False

        self._stop_flag.clear()
        self.is_running = True

        auto_tests = [tc for tc in FOTA_TEST_CASES
                       if tc.auto_func and tc.default_result != 'NA']
        self.progress = {
            'current_test': 0,
            'total_tests': len(auto_tests),
            'phase': 'running',
            'status': 'running'
        }

        self._thread = threading.Thread(
            target=self._run_all_worker,
            args=(serial, dut2_serial, dut2_phone, auto_tests),
            daemon=True
        )
        self._thread.start()
        return True

    def stop(self):
        self._stop_flag.set()

    def set_manual_result(self, test_id: int, result: str, remark: str = "") -> bool:
        """Marca resultado manual para un test"""
        if test_id not in self.results:
            return False
        self.results[test_id].update({
            'result': result,
            'remark': remark,
            'timestamp': datetime.now().isoformat()
        })
        self._log(f"Test #{test_id} marcado manualmente: {result}")
        return True

    def generate_report(self, model: str, sw_version: str, tester: str, sp_date: str = "") -> str:
        """Genera informe JSON de resultados"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = os.path.join('data', 'fota_reports')
        os.makedirs(report_dir, exist_ok=True)

        report = {
            'header': {
                'model': model,
                'sw_version': sw_version,
                'tester': tester,
                'sp_date': sp_date,
                'date': datetime.now().isoformat(),
            },
            'summary': {
                'total': len(FOTA_TEST_CASES),
                'pass': sum(1 for r in self.results.values() if r['result'] == 'Pass'),
                'fail': sum(1 for r in self.results.values() if r['result'] == 'Fail'),
                'na': sum(1 for r in self.results.values() if r['result'] == 'NA'),
                'pending': sum(1 for r in self.results.values() if r['result'] == 'pending'),
            },
            'tests': []
        }

        for tc in FOTA_TEST_CASES:
            r = self.results[tc.id]
            report['tests'].append({
                'id': tc.id,
                'title': tc.title,
                'description': tc.description,
                'expected': tc.expected,
                'result': r['result'],
                'remark': r['remark'],
                'automation': tc.automation,
            })

        filepath = os.path.join(report_dir, f'fota_report_{timestamp}.json')
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

        self._log(f"Informe generado: {filepath}")
        return filepath

    def generate_excel_report(self, model: str, sw_version: str, tester: str, sp_date: str = "") -> str:
        """Genera informe Excel con formato similar al original"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            self._log("openpyxl no disponible, generando solo JSON", "ERROR")
            return self.generate_report(model, sw_version, tester, sp_date)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_dir = os.path.join('data', 'fota_reports')
        os.makedirs(report_dir, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FOTA"

        # Estilos
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=10)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        na_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        # Header info
        ws['C2'] = "Test Case FOTA Claro"
        ws['C2'].font = Font(bold=True, size=14)
        ws['C3'] = "Model"
        ws['D3'] = model
        ws['C4'] = "SW Version"
        ws['D4'] = sw_version
        ws['C5'] = "Date"
        ws['D5'] = datetime.now().strftime("%Y-%m-%d")
        ws['C6'] = "Tester"
        ws['D6'] = tester
        ws['C7'] = "SP Date"
        ws['D7'] = sp_date

        # Contadores
        pass_count = sum(1 for r in self.results.values() if r['result'] == 'Pass')
        fail_count = sum(1 for r in self.results.values() if r['result'] == 'Fail')
        na_count = sum(1 for r in self.results.values() if r['result'] == 'NA')
        pend_count = sum(1 for r in self.results.values() if r['result'] == 'pending')

        ws['G2'] = "Result"
        ws['H2'] = "Count"
        ws['G3'] = "Pass"
        ws['H3'] = pass_count
        ws['G3'].fill = pass_fill
        ws['G4'] = "Fail"
        ws['H4'] = fail_count
        ws['G4'].fill = fail_fill
        ws['G5'] = "NA"
        ws['H5'] = na_count
        ws['G5'].fill = na_fill
        ws['G6'] = "Pending"
        ws['H6'] = pend_count

        # Encabezados de tabla
        row = 10
        headers = ['#', 'Title', 'Description', 'Expected', 'Result', 'Remark']
        cols = ['B', 'C', 'D', 'E', 'F', 'G']
        for col, header in zip(cols, headers):
            cell = ws[f'{col}{row}']
            cell.value = header
            cell.font = title_font
            cell.border = thin_border

        # Test cases
        for tc in FOTA_TEST_CASES:
            row += 1
            r = self.results[tc.id]
            ws[f'B{row}'] = tc.id
            ws[f'C{row}'] = tc.title
            ws[f'D{row}'] = tc.description
            ws[f'E{row}'] = tc.expected
            ws[f'F{row}'] = r['result']
            ws[f'G{row}'] = r['remark']

            # Colorear resultado
            result_cell = ws[f'F{row}']
            if r['result'] == 'Pass':
                result_cell.fill = pass_fill
            elif r['result'] == 'Fail':
                result_cell.fill = fail_fill
            elif r['result'] == 'NA':
                result_cell.fill = na_fill

            for col in cols:
                ws[f'{col}{row}'].border = thin_border
                ws[f'{col}{row}'].alignment = Alignment(wrap_text=True, vertical='top')

        # Ajustar anchos
        ws.column_dimensions['B'].width = 5
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 35
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 25

        filepath = os.path.join(report_dir, f'FOTA_Report_{model}_{timestamp}.xlsx')
        wb.save(filepath)
        self._log(f"Informe Excel generado: {filepath}")
        return filepath

    # ==================== WORKER THREAD ====================

    def _run_all_worker(self, serial, dut2_serial, dut2_phone, auto_tests):
        """Thread worker que ejecuta todos los tests automáticos"""
        try:
            self._log(f"Iniciando ejecución automática: {len(auto_tests)} tests")
            for idx, tc in enumerate(auto_tests):
                if self._stop_flag.is_set():
                    self._log("Ejecución detenida por el usuario")
                    break

                with self._lock:
                    self.progress['current_test'] = idx + 1
                    self.progress['phase'] = f"Test #{tc.id}: {tc.title}"

                self.run_single(serial, tc.id, dut2_serial, dut2_phone)

            with self._lock:
                self.progress['status'] = 'completed'
                self.progress['phase'] = 'done'
            self._log("Ejecución automática completada")

        except Exception as e:
            self._log(f"Error en ejecución automática: {e}", "ERROR")
            with self._lock:
                self.progress['status'] = 'error'
        finally:
            self.is_running = False

    # ==================== FUNCIONES DE TEST ====================

    def _take_screenshot(self, serial: str, test_id: int) -> Optional[str]:
        """Captura screenshot como evidencia de un test"""
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"screenshots/fota_test{test_id}_{ts}.png"
        ok, path = self.adb.capture_screenshot(serial, filename)
        return path if ok else None

    def _test_sw_version(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #1: Verificar SW Version"""
        info = self.adb.get_sw_version(serial)
        display = info.get('display_id', '')
        incremental = info.get('incremental', '')
        security = info.get('security_patch', '')
        model = info.get('model', '')

        screenshot = self._take_screenshot(serial, 1)

        if display and incremental:
            remark = f"Model: {model}\nBuild: {display}\nIncremental: {incremental}\nSP: {security}"
            return {'result': 'Pass', 'remark': remark, 'screenshot': screenshot}
        return {'result': 'Fail', 'remark': 'No se pudo obtener versión de SW'}

    def _test_network_name(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #2: Verificar nombre de red"""
        device = self.adb.refresh_device(serial)
        if not device:
            return {'result': 'Fail', 'remark': 'Dispositivo no encontrado'}

        operator = device.sim_operator
        network = device.network_type
        screenshot = self._take_screenshot(serial, 2)

        if operator and operator != 'Sin operador':
            return {
                'result': 'Pass',
                'remark': f"Operador: {operator}, Red: {network}",
                'screenshot': screenshot
            }
        return {'result': 'Fail', 'remark': 'No se detectó operador', 'screenshot': screenshot}

    def _test_voicemail(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #4: Voice mail (semi - lanza buzón, tester confirma)"""
        # Marcar *611 o #611 (buzón Claro)
        ok, msg = self.adb.make_call(serial, "*611")
        time.sleep(5)
        screenshot = self._take_screenshot(serial, 4)
        self.adb.end_call(serial)
        return {
            'result': 'pending',
            'remark': f"Buzón marcado. Verificar manualmente. {msg}",
            'screenshot': screenshot
        }

    def _test_mms(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #7: MMS (semi - abre app mensajes)"""
        self.adb.run_command(
            "shell am start -a android.intent.action.SEND -t image/* --es sms_body 'Test MMS FOTA'",
            serial
        )
        time.sleep(3)
        screenshot = self._take_screenshot(serial, 7)
        return {
            'result': 'pending',
            'remark': 'App de mensajes abierta. Enviar MMS manualmente y verificar.',
            'screenshot': screenshot
        }

    def _test_sms(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #8: SMS - enviar y verificar recepción"""
        if not dut2_serial or not dut2_phone:
            return {'result': 'pending', 'remark': 'Se requiere DUT2 para prueba de SMS'}

        test_msg = f"Test FOTA SMS {datetime.now().strftime('%H:%M:%S')}"
        # Obtener número de DUT1
        dut1_phone = self.adb._get_phone_number(serial)

        # Enviar SMS de DUT1 a DUT2
        ok1, msg1 = self.adb.send_sms(serial, dut2_phone, test_msg)
        if not ok1:
            return {'result': 'Fail', 'remark': f'Error enviando SMS: {msg1}'}

        time.sleep(5)

        # Verificar recepción en DUT2
        ok2, msg2 = self.adb.check_sms_received(dut2_serial, dut1_phone or serial, test_msg, timeout=30)

        # Enviar SMS de vuelta (DUT2 a DUT1)
        if dut1_phone:
            ok3, msg3 = self.adb.send_sms(dut2_serial, dut1_phone, f"Reply: {test_msg}")
            time.sleep(5)
            ok4, msg4 = self.adb.check_sms_received(serial, dut2_phone, f"Reply: {test_msg}", timeout=30)
        else:
            ok3, ok4 = True, True

        screenshot = self._take_screenshot(serial, 8)

        if ok1 and ok2:
            return {'result': 'Pass', 'remark': 'SMS enviado y recibido correctamente', 'screenshot': screenshot}
        return {'result': 'Fail', 'remark': f'Envío: {msg1}. Recepción: {msg2}', 'screenshot': screenshot}

    def _test_browser(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #9: Browser (semi)"""
        self.adb.run_command(
            "shell am start -a android.intent.action.VIEW -d https://www.google.com",
            serial
        )
        time.sleep(8)
        screenshot = self._take_screenshot(serial, 9)
        return {
            'result': 'pending',
            'remark': 'Browser abierto con Google. Verificar que carga correctamente.',
            'screenshot': screenshot
        }

    def _test_stk(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #10: SIM/STK (semi)"""
        self.adb.run_command(
            "shell am start -n com.android.stk/.StkLauncherActivity",
            serial
        )
        time.sleep(3)
        screenshot = self._take_screenshot(serial, 10)
        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'pending',
            'remark': 'STK abierto. Verificar nombre y opciones.',
            'screenshot': screenshot
        }

    def _test_idle(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #12: Idle screen (semi)"""
        # Ir a home y capturar
        self.adb.run_command("shell input keyevent KEYCODE_HOME", serial)
        time.sleep(2)
        screenshot = self._take_screenshot(serial, 12)
        return {
            'result': 'pending',
            'remark': 'Screenshot de idle. Verificar: barra indicadora, RSSI, wallpaper, hora/fecha, batería.',
            'screenshot': screenshot
        }

    def _test_menu_tree(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #13: Menu tree (semi)"""
        # Abrir settings
        self.adb.run_command("shell am start -a android.settings.SETTINGS", serial)
        time.sleep(3)
        screenshot = self._take_screenshot(serial, 13)
        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'pending',
            'remark': 'Settings abiertos. Verificar navegación sin errores ni traducciones incorrectas.',
            'screenshot': screenshot
        }

    def _test_power_onoff(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #14: Power ON/OFF"""
        self._log("Reiniciando dispositivo...")
        ok, msg = self.adb.reboot_and_wait(serial, timeout=120)
        if ok:
            time.sleep(5)
            screenshot = self._take_screenshot(serial, 14)
            return {'result': 'Pass', 'remark': 'Reinicio exitoso', 'screenshot': screenshot}
        return {'result': 'Fail', 'remark': msg}

    def _test_volte_icon(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #15: VoLTE icon visible"""
        device = self.adb.refresh_device(serial)
        if not device:
            return {'result': 'Fail', 'remark': 'Dispositivo no encontrado'}

        screenshot = self._take_screenshot(serial, 15)

        if device.volte_enabled:
            return {'result': 'Pass', 'remark': 'VoLTE habilitado', 'screenshot': screenshot}
        return {'result': 'Fail', 'remark': 'VoLTE no detectado', 'screenshot': screenshot}

    def _test_volte_airplane(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #16: VoLTE persiste después de airplane mode"""
        # Verificar VoLTE antes
        device = self.adb.refresh_device(serial)
        volte_before = device.volte_enabled if device else False

        # Activar modo avión
        self.adb.set_airplane_mode(serial, True)
        time.sleep(3)

        # Verificar VoLTE durante avión (debe desaparecer)
        device_air = self.adb.refresh_device(serial)
        volte_airplane = device_air.volte_enabled if device_air else False

        # Desactivar modo avión
        self.adb.set_airplane_mode(serial, False)
        time.sleep(10)  # Esperar reconexión

        # Verificar VoLTE después
        device_after = self.adb.refresh_device(serial)
        volte_after = device_after.volte_enabled if device_after else False
        screenshot = self._take_screenshot(serial, 16)

        if volte_before and not volte_airplane and volte_after:
            return {'result': 'Pass',
                    'remark': f'VoLTE antes={volte_before}, avión={volte_airplane}, después={volte_after}',
                    'screenshot': screenshot}
        return {'result': 'Fail',
                'remark': f'VoLTE antes={volte_before}, avión={volte_airplane}, después={volte_after}',
                'screenshot': screenshot}

    def _test_volte_call(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #17: VoLTE call (semi - hace llamada, tester confirma calidad)"""
        if not dut2_serial or not dut2_phone:
            return {'result': 'pending', 'remark': 'Se requiere DUT2 para llamada VoLTE'}

        ok, msg = self.adb.make_call(serial, dut2_phone)
        if not ok:
            return {'result': 'Fail', 'remark': f'Error haciendo llamada: {msg}'}

        time.sleep(5)
        # Intentar contestar en DUT2
        self.adb.answer_call(dut2_serial)
        time.sleep(3)

        screenshot = self._take_screenshot(serial, 17)

        # Mantener 10 segundos
        time.sleep(10)
        self.adb.end_call(serial)
        self.adb.end_call(dut2_serial)

        return {
            'result': 'pending',
            'remark': 'Llamada VoLTE realizada 10s. Verificar calidad de voz.',
            'screenshot': screenshot
        }

    def _test_volte_conference(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #19: VoLTE conference (semi)"""
        return {
            'result': 'pending',
            'remark': 'Conferencia VoLTE requiere 3 dispositivos. Ejecutar manualmente.',
        }

    def _test_volte_conference_b_ends(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #20: VoLTE conference B ends (semi)"""
        return {
            'result': 'pending',
            'remark': 'Conferencia VoLTE (B cuelga) requiere 3 dispositivos. Ejecutar manualmente.',
        }

    def _test_volte_sms(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #22: SMS durante VoLTE activo"""
        # Verificar VoLTE
        device = self.adb.refresh_device(serial)
        if not device or not device.volte_enabled:
            return {'result': 'Fail', 'remark': 'VoLTE no está activo'}

        if not dut2_serial or not dut2_phone:
            return {'result': 'pending', 'remark': 'Se requiere DUT2'}

        # Enviar SMS
        test_msg = f"VoLTE SMS Test {datetime.now().strftime('%H:%M:%S')}"
        ok, msg = self.adb.send_sms(serial, dut2_phone, test_msg)
        screenshot = self._take_screenshot(serial, 22)

        if ok:
            return {'result': 'Pass', 'remark': 'SMS enviado con VoLTE activo', 'screenshot': screenshot}
        return {'result': 'Fail', 'remark': f'Error: {msg}', 'screenshot': screenshot}

    def _test_vowifi_registration(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #23: VoWiFi registration"""
        device = self.adb.refresh_device(serial)
        screenshot = self._take_screenshot(serial, 23)

        if device and device.vowifi_enabled:
            return {'result': 'Pass', 'remark': 'VoWiFi registrado', 'screenshot': screenshot}
        return {
            'result': 'pending',
            'remark': f'VoWiFi: {device.vowifi_enabled if device else "N/A"}. Verificar ícono manualmente.',
            'screenshot': screenshot
        }

    def _test_vowifi_airplane(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #24: VoWiFi en modo avión + WiFi"""
        # Activar modo avión
        self.adb.set_airplane_mode(serial, True)
        time.sleep(3)

        # Habilitar WiFi
        self.adb.set_wifi(serial, True)
        time.sleep(10)  # Esperar conexión WiFi + registro IMS

        device = self.adb.refresh_device(serial)
        screenshot = self._take_screenshot(serial, 24)

        # Restaurar
        self.adb.set_airplane_mode(serial, False)
        time.sleep(5)

        wifi_connected, ssid = self.adb.check_wifi_connected(serial)

        if wifi_connected:
            return {
                'result': 'Pass',
                'remark': f'WiFi conectado en modo avión. SSID: {ssid}',
                'screenshot': screenshot
            }
        return {
            'result': 'pending',
            'remark': 'Verificar registro IMS manualmente en modo avión + WiFi',
            'screenshot': screenshot
        }

    def _test_vowifi_call(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #25: VoWiFi call (semi)"""
        if not dut2_phone:
            return {'result': 'pending', 'remark': 'Se requiere DUT2'}

        ok, msg = self.adb.make_call(serial, dut2_phone)
        time.sleep(5)
        if dut2_serial:
            self.adb.answer_call(dut2_serial)
        time.sleep(10)
        screenshot = self._take_screenshot(serial, 25)
        self.adb.end_call(serial)
        if dut2_serial:
            self.adb.end_call(dut2_serial)

        return {
            'result': 'pending',
            'remark': 'Llamada VoWiFi realizada. Verificar calidad de voz.',
            'screenshot': screenshot
        }

    def _test_call_types(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #28: Diferentes tipos de llamada (semi)"""
        if not dut2_phone:
            return {'result': 'pending', 'remark': 'Se requiere DUT2'}

        # MO call
        ok, _ = self.adb.make_call(serial, dut2_phone)
        time.sleep(5)
        if dut2_serial:
            self.adb.answer_call(dut2_serial)
        time.sleep(5)
        self.adb.end_call(serial)
        time.sleep(3)

        screenshot = self._take_screenshot(serial, 28)
        return {
            'result': 'pending',
            'remark': 'MO call realizada. Verificar historial de llamadas manualmente.',
            'screenshot': screenshot
        }

    def _test_multiparty_mo(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #29: Multiparty MO (semi)"""
        return {
            'result': 'pending',
            'remark': 'Multiparty MO requiere múltiples dispositivos. Ejecutar manualmente.'
        }

    def _test_multiparty_mt(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #30: Multiparty MT (semi)"""
        return {
            'result': 'pending',
            'remark': 'Multiparty MT requiere múltiples dispositivos. Ejecutar manualmente.'
        }

    def _test_wifi(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #32: WiFi connection"""
        # Habilitar WiFi
        self.adb.set_wifi(serial, True)
        time.sleep(5)

        connected, ssid = self.adb.check_wifi_connected(serial)
        screenshot = self._take_screenshot(serial, 32)

        if connected:
            # Desconectar y reconectar
            self.adb.set_wifi(serial, False)
            time.sleep(3)
            self.adb.set_wifi(serial, True)
            time.sleep(8)
            connected2, ssid2 = self.adb.check_wifi_connected(serial)

            if connected2:
                return {
                    'result': 'Pass',
                    'remark': f'WiFi conectado: {ssid}. Reconexión OK: {ssid2}',
                    'screenshot': screenshot
                }
            return {'result': 'Fail', 'remark': f'WiFi conectó pero no reconectó', 'screenshot': screenshot}

        return {'result': 'Fail', 'remark': 'WiFi no se conectó', 'screenshot': screenshot}

    def _test_wifi_streaming(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #33: WiFi streaming (semi)"""
        self.adb.run_command(
            "shell am start -a android.intent.action.VIEW -d https://www.youtube.com",
            serial
        )
        time.sleep(8)
        screenshot = self._take_screenshot(serial, 33)
        return {
            'result': 'pending',
            'remark': 'YouTube abierto. Verificar que reproduce video correctamente.',
            'screenshot': screenshot
        }

    def _test_bluetooth_search(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #34: Bluetooth search"""
        found, devices = self.adb.scan_bluetooth_devices(serial)
        screenshot = self._take_screenshot(serial, 34)

        if found:
            device_list = ", ".join(devices[:5])
            return {
                'result': 'Pass',
                'remark': f'BT encontró {len(devices)} dispositivos: {device_list}',
                'screenshot': screenshot
            }
        return {
            'result': 'Fail',
            'remark': 'Bluetooth no encontró dispositivos',
            'screenshot': screenshot
        }

    def _test_alarm(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #36: Alarm"""
        # Abrir app de reloj/alarma
        self.adb.run_command(
            "shell am start -a android.intent.action.SET_ALARM "
            "--ei android.intent.extra.alarm.HOUR 12 "
            "--ei android.intent.extra.alarm.MINUTES 0 "
            "--es android.intent.extra.alarm.MESSAGE 'FOTA Test'",
            serial
        )
        time.sleep(3)
        screenshot = self._take_screenshot(serial, 36)
        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'Pass',
            'remark': 'Alarma configurada via intent',
            'screenshot': screenshot
        }

    def _test_calculator(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #37: Calculator"""
        self.adb.launch_app(serial, "com.google.android.calculator")
        time.sleep(3)
        screenshot = self._take_screenshot(serial, 37)
        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'Pass',
            'remark': 'Calculadora abierta correctamente',
            'screenshot': screenshot
        }

    def _test_camera(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #38: Camera"""
        # Abrir cámara
        self.adb.run_command(
            "shell am start -a android.media.action.STILL_IMAGE_CAMERA",
            serial
        )
        time.sleep(4)
        screenshot = self._take_screenshot(serial, 38)

        # Capturar foto via keyevent
        self.adb.run_command("shell input keyevent KEYCODE_CAMERA", serial)
        time.sleep(3)

        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'Pass',
            'remark': 'Cámara abierta y foto capturada',
            'screenshot': screenshot
        }

    def _test_save_contact(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #41: Save contact"""
        ok, msg = self.adb.insert_contact(serial, "FOTA Test Contact", "+5731234567")
        screenshot = self._take_screenshot(serial, 41)

        if ok:
            return {'result': 'Pass', 'remark': msg, 'screenshot': screenshot}
        return {'result': 'Fail', 'remark': msg, 'screenshot': screenshot}

    def _test_customization(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #43: Customization - verificar apps preinstaladas"""
        # Apps comunes de Claro/AMX
        apps_to_check = [
            'com.claro.miclaro',
            'com.claro.clarovideo',
            'com.claro.claromusica',
        ]

        found = []
        not_found = []
        for pkg in apps_to_check:
            ver = self.adb.get_installed_app_version(serial, pkg)
            if ver:
                found.append(f"{pkg.split('.')[-1]}: v{ver}")
            else:
                not_found.append(pkg.split('.')[-1])

        screenshot = self._take_screenshot(serial, 43)
        remark = f"Encontradas: {', '.join(found) if found else 'ninguna'}"
        if not_found:
            remark += f"\nNo encontradas: {', '.join(not_found)}"

        return {
            'result': 'Pass' if found else 'pending',
            'remark': remark,
            'screenshot': screenshot
        }

    def _test_sysdll(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #44: SYSDLL app version"""
        # Buscar app SYSDLL (varía por fabricante)
        possible_pkgs = [
            'com.sysdll.app',
            'com.amt.sysdll',
            'com.amx.sysdll',
        ]

        for pkg in possible_pkgs:
            ver = self.adb.get_installed_app_version(serial, pkg)
            if ver:
                return {
                    'result': 'Pass',
                    'remark': f'SYSDLL encontrada: {pkg} v{ver}'
                }

        return {
            'result': 'pending',
            'remark': 'SYSDLL no encontrada con packages conocidos. Verificar manualmente.'
        }

    def _test_fota_preservation(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #45: FOTA preservation (semi)"""
        screenshot = self._take_screenshot(serial, 45)
        return {
            'result': 'pending',
            'remark': 'Captura post-FOTA. Verificar que wallpaper, íconos, ringtone y apps se mantienen.',
            'screenshot': screenshot
        }

    def _test_fota_popup(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #46: FOTA popup (semi)"""
        screenshot = self._take_screenshot(serial, 46)
        return {
            'result': 'pending',
            'remark': 'Verificar que popup FOTA se muestra en idioma local.',
            'screenshot': screenshot
        }

    def _test_fota_sw_version(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #47: FOTA SW version post-update"""
        info = self.adb.get_sw_version(serial)
        display = info.get('display_id', '')
        screenshot = self._take_screenshot(serial, 47)

        if display:
            return {
                'result': 'Pass',
                'remark': f'SW post-FOTA: {display}',
                'screenshot': screenshot
            }
        return {'result': 'Fail', 'remark': 'No se pudo leer SW version'}

    def _test_fota_wordings(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #48: FOTA wordings (semi)"""
        # Abrir actualización de sistema
        self.adb.run_command(
            "shell am start -a android.settings.SYSTEM_UPDATE_SETTINGS",
            serial
        )
        time.sleep(5)
        screenshot = self._take_screenshot(serial, 48)
        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'pending',
            'remark': 'Pantalla de actualización abierta. Verificar textos en español.',
            'screenshot': screenshot
        }

    def _test_fota_auto(self, serial, dut2_serial=None, dut2_phone=None) -> Dict:
        """Test #50: FOTA solo descarga por WiFi (semi)"""
        self.adb.run_command(
            "shell am start -a android.settings.SYSTEM_UPDATE_SETTINGS",
            serial
        )
        time.sleep(5)
        screenshot = self._take_screenshot(serial, 50)
        self.adb.run_command("shell input keyevent 4", serial)
        return {
            'result': 'pending',
            'remark': 'Verificar que FOTA no descarga por datos móviles, solo por WiFi.',
            'screenshot': screenshot
        }
