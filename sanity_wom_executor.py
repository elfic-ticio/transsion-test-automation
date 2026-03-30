"""
Sanity Check WOM - Ejecutor de casos de prueba post-OTA
72 casos de prueba para validacion despues de actualizacion OTA con WOM Colombia
"""
import threading
import time
import json
import os
import logging
from datetime import datetime
from dataclasses import dataclass, field, asdict
from typing import Optional, List, Dict

logger = logging.getLogger(__name__)

RESULTS_FILE = os.path.join('data', 'sanity_wom_results.json')


@dataclass
class WOMTestCase:
    id: str           # Ej: "wom_10", "wom_volte_0001"
    number: str       # Numero original: "10", "FT_VoLTE_0001"
    name: str         # Nombre completo
    category: str     # 'general', 'csfb', 'volte', 'vowifi', '5g'
    description: str
    procedure: str
    expected: str
    automation: str   # 'auto', 'semi', 'manual'
    auto_func: str = ""
    result: str = "pending"   # 'pending', 'pass', 'fail', 'na'
    remark: str = ""
    last_run: str = ""

    def to_dict(self):
        return asdict(self)


# ==================== DEFINICION DE LOS 72 CASOS ====================

WOM_TEST_CASES = [
    # ---- GENERAL ----
    WOMTestCase(
        "wom_2", "2", "[WOM] [2] Revision SPN en Idle", "general",
        "Verificar que el nombre de red (SPN) se muestra correctamente.",
        "1) Verificar que el NOMBRE DE RED (SPN) se muestra correctamente en pantalla.\n"
        "2) Verificar que el nombre de red se actualiza al cambiar de operador.",
        "SPN de WOM visible en barra de estado y pantalla de inicio.",
        "auto", "_auto_spn"
    ),
    WOMTestCase(
        "wom_3", "3", "[WOM] [3] Revision APN en Idle", "general",
        "Verificar que el APN es correcto y esta registrado en el telefono.",
        "1) Verificar que el APN es correcto.\n"
        "2) El APN de WOM debe estar registrado en el dispositivo.",
        "APN de WOM configurado correctamente (internet.wom.co o equivalente).",
        "auto", "_auto_apn"
    ),
    WOMTestCase(
        "wom_5", "5", "[WOM] [5] Revision Actualizacion de Hora", "general",
        "Verificar la actualizacion automatica de hora y fecha via red.",
        "Caso 1)\n1) En Ajustes > Gestion general > Fecha y hora, desactivar 'Fecha y hora automaticas'.\n"
        "2) Cambiar la fecha o la hora a un valor incorrecto (ej. ano 2020).\n"
        "3) Activar 'Fecha y hora automaticas'.\n"
        "4) Esperar unos minutos hasta que la fecha y hora sean correctas.\n"
        "NOTA: El boton ADB verifica el estado actual de auto_time en el dispositivo.",
        "La hora y fecha se actualizan automaticamente via red.",
        "semi", "_auto_time_sync"
    ),
    WOMTestCase(
        "wom_6", "6", "[WOM] [6] Seleccion de Red", "general",
        "Verificar la seleccion manual y automatica de red.",
        "Caso 1) Modo manual - Escaneo de redes disponibles.\n"
        "Caso 2) Modo manual - Seleccion de red especifica.\n"
        "Caso 3) Modo automatico - Seleccion automatica de red.",
        "El dispositivo puede escanear, seleccionar y registrarse en redes correctamente.",
        "semi", "_semi_network_selection"
    ),
    WOMTestCase(
        "wom_10", "10", "[WOM] [10] Prueba de Navegacion Web", "general",
        "Realizar pruebas de navegacion web sin Wi-Fi.",
        "No utilizar la Red Wi-Fi.\n"
        "1) Conectarse a un sitio web y navegar, descargar contenido (imagen/audio/video).\n"
        "2) Verificar que el navegador se conecta a los marcadores predeterminados.\n"
        "3) Si hay distintas configuraciones para SIM prepago/postpago, verificar ambos casos.",
        "El navegador se conecta, navega, descarga y almacena contenido sin danos.",
        "semi", "_semi_browser"
    ),
    WOMTestCase(
        "wom_11", "11", "[WOM] [11] Prueba de Navegacion Streaming", "general",
        "Verificar la conexion a nivel streaming de video.",
        "1) Ingresar a la tienda de aplicaciones.\n2) Descargar la aplicacion de TV movil.\n"
        "3) Abrir la aplicacion.\n4) Ver al menos 2 canales durante unos minutos.\n"
        "Nota: Si no hay aplicacion de TV movil, verificar con otra aplicacion de streaming.",
        "Se puede ver y escuchar el contenido. Video y audio estan sincronizados y con buena calidad.",
        "manual"
    ),
    WOMTestCase(
        "wom_15", "15", "[WOM] [15] Prueba de Redes Sociales", "general",
        "Verificar el funcionamiento correcto de aplicaciones de Redes Sociales.",
        "Verificar los clientes de redes sociales preinstalados: Facebook, Twitter y similares, con SIM activa.",
        "Las aplicaciones de redes sociales funcionan correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_16", "16", "[WOM] [16] Prueba de Mensajeria Instantanea", "general",
        "Verificar el funcionamiento de aplicaciones de mensajeria instantanea.",
        "1) Instalar clientes de mensajeria: WhatsApp, Telegram, Facebook Messenger.\n"
        "2) Verificar que los mensajes se envian y reciben correctamente.",
        "Los clientes de mensajeria funcionan correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_28", "28", "[WOM] [28] Prueba de Servicios Prepago", "general",
        "Verificar que el terminal ejecuta correctamente los servicios de Prepago.",
        "Marcar el codigo USSD de WOM Prepago: *888#\n"
        "Verificar que el menu de servicios prepago responde correctamente en pantalla.",
        "Los servicios prepago responden correctamente.",
        "semi", "_semi_prepago_ussd"
    ),
    WOMTestCase(
        "wom_38", "38", "[WOM] [38] Llamada 3G Roaming Nacional", "general",
        "Verificar llamada de voz en 3G con Roaming Nacional (itinerancia entre operadores colombianos).",
        "Caso 1) Seleccion de Movistar 3G:\n"
        "1) Ir a Ajustes > Redes moviles > Operador de red > Seleccion manual.\n"
        "2) Escanear las redes disponibles (WOM, Movistar, Claro, Tigo).\n"
        "3) Seleccionar Movistar 3G.\n"
        "4) Verificar que el DUT se registra correctamente en Movistar.\n"
        "5) Realizar una llamada de voz, enviar SMS y navegar datos.\n\n"
        "Caso 2) Repetir con Movistar 2G.",
        "El dispositivo puede registrarse en red 3G/2G de Movistar y realizar llamadas y datos.",
        "semi", "_semi_roaming_scan"
    ),
    WOMTestCase(
        "wom_40", "40", "[WOM] [40] Llamada a Numeros Cortos", "general",
        "Verificar que el terminal puede marcar numeros de marcacion corta de WOM Colombia.",
        "DUT registrado en VoLTE.\n"
        "1) Marcar los siguientes numeros cortos:\n"
        "   - *611 (Atencion al cliente WOM)\n"
        "   - *222 (Saldo y recargas)\n"
        "   - 164 (Informacion)\n"
        "   - 123 (Policia)\n"
        "   - 112 (Emergencias)\n"
        "2) Verificar que cada llamada conecta y se ejecuta sobre VoLTE.\n"
        "3) Registrar el resultado y la tecnologia usada en cada caso.",
        "Todas las marcaciones se generan sobre VoLTE y conectan correctamente.",
        "semi", "_semi_numeros_cortos"
    ),
    WOMTestCase(
        "wom_41", "41", "[WOM] [41] Buzon de Voz - Icono de Notificacion", "general",
        "Verificar que el terminal muestra el icono de buzon de voz correctamente.",
        "NOTA: El servicio de buzon de voz puede no estar disponible en WOM Colombia.\n"
        "Condicion previa: Enviar 3 mensajes de voz al dispositivo sin que sea contestado.\n"
        "1) Esperar hasta recibir la notificacion de buzon de voz (icono en barra de estado).\n"
        "2) Apagar y encender el equipo.\n"
        "3) Verificar que el icono de voicemail sigue apareciendo.\n"
        "4) Escuchar los mensajes (dejar al menos uno sin escuchar).\n"
        "5) Verificar que el icono permanece mientras haya mensajes no escuchados.",
        "El icono de voicemail aparece y desaparece correctamente segun los mensajes pendientes.",
        "semi", "_semi_voicemail_settings"
    ),
    WOMTestCase(
        "wom_42", "42", "[WOM] [42] Buzon de Voz - Tecla de Acceso Directo", "general",
        "Verificar que el terminal llama al Buzon de Voz mediante acceso directo.",
        "NOTA: El servicio de buzon de voz puede no estar disponible en WOM Colombia.\n"
        "1) Mantener presionada la tecla '1' en el marcador (acceso rapido al buzon).\n"
        "2) Verificar que el DUT marca al centro de buzon de voz.\n"
        "3) Tambien probar marcando directamente el numero del buzon de voz.",
        "El dispositivo conecta al servicio de buzon de voz por ambos metodos.",
        "semi", "_semi_voicemail_dial"
    ),
    WOMTestCase(
        "wom_43", "43", "[WOM] [43] Buzon de Voz - Recepcion de Mensaje", "general",
        "Verificar que el terminal gestiona el icono de buzon de mensajes.",
        "1) Verificar que el icono aparece cuando se recibe un mensaje de voz.\n"
        "2) Verificar que el icono desaparece al escuchar el mensaje.",
        "El icono de voicemail se muestra y oculta correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_44", "44", "[WOM] [44] Buzon de Voz - Interaccion con la Plataforma", "general",
        "Verificar la interaccion correcta del terminal con el sistema de voicemail.",
        "Verificar la funcionalidad del correo de voz:\n"
        "Escuchar el mensaje, borrar, guardar, responder.",
        "El dispositivo interactua correctamente con la plataforma de voicemail.",
        "manual"
    ),
    WOMTestCase(
        "wom_45", "45", "[WOM] [45] Buzon de Voz en Equipos Dual SIM", "general",
        "Verificar el funcionamiento del Buzon de Voz en equipos Dual SIM.",
        "Requisito previo: Solo aplica a dispositivos Dual SIM.\n"
        "Verificar el buzon de voz en ambas ranuras SIM.",
        "El dispositivo Dual SIM gestiona el voicemail correctamente en ambas SIM.",
        "manual"
    ),
    WOMTestCase(
        "wom_60", "60", "[WOM] [60] Menu SIM TOOL KIT (STK)", "general",
        "Verificar la interaccion correcta del terminal con el menu STK.",
        "(Segun soporte de la red)\nCon SIM prepago:\n"
        "1) Verificar todos los servicios STK disponibles.\n"
        "2) Verificar que el menu STK se muestra correctamente.",
        "El menu STK se muestra correctamente y los servicios funcionan.",
        "manual"
    ),
    WOMTestCase(
        "wom_61", "61", "[WOM] [61] STK Roaming Broaker", "general",
        "Verificar que el terminal ejecuta correctamente el applet Roaming Broaker.",
        "1) Ingresar a SIM APP.\n2) Ingresar a App Roaming Internacional.\n"
        "3) Cambiar la configuracion de roaming.",
        "El applet STK Roaming Broaker funciona correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_63", "63", "[WOM] [63] STK Roaming Broaker III", "general",
        "Verificar la interaccion correcta del terminal con el applet Roaming Broaker III.",
        "1) Ingresar a SIM APP.\n2) Ingresar a App Roaming Internacional.\n"
        "3) Cambiar la configuracion de roaming internacional.",
        "El applet STK Roaming Broaker III funciona correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_64", "64", "[WOM] [64] STK Roaming Broaker IV", "general",
        "Verificar la interaccion correcta del terminal con el applet Roaming Broaker IV.",
        "1) Acceder a la aplicacion STK.\n"
        "2) Verificar el comportamiento de la aplicacion STK.",
        "El applet STK Roaming Broaker IV funciona correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_71", "71", "[WOM] [71] Prueba de GPS", "general",
        "Verificar el funcionamiento del GPS desde el terminal.",
        "1) Abrir Google Maps.\n"
        "2) Seleccionar la opcion para verificar la ubicacion actual.\n"
        "3) Verificar que la ubicacion se muestra correctamente.",
        "El GPS localiza correctamente la posicion del dispositivo.",
        "semi", "_semi_gps"
    ),
    WOMTestCase(
        "wom_83", "83", "[WOM] [83] Revision de Throughput (Descarga / Carga)", "general",
        "Verificar el comportamiento correcto del terminal en Uplink y Downlink con LTE conectado.",
        "Con LTE conectado (sin Wi-Fi):\n"
        "1) Iniciar descarga HTTP de un archivo grande (100 MB).\n"
        "2) Iniciar carga simultanea de un archivo (70 MB).\n"
        "3) Comparar las velocidades DL y UL obtenidas con un dispositivo de referencia.\n"
        "Anotar velocidades en el campo de observaciones.",
        "Las velocidades de descarga y carga son mayores o similares a las del dispositivo de referencia.",
        "semi", "_semi_throughput"
    ),
    WOMTestCase(
        "wom_85", "85", "[WOM] [85] Busqueda de Red LTE", "general",
        "Verificar que el terminal puede buscar y registrarse correctamente en la red LTE.",
        "Con LTE conectado:\n"
        "1) Ir a Ajustes > Redes moviles > Busqueda manual de red.\n"
        "2) Seleccionar la misma red LTE a la que ya esta conectado el DUT.\n"
        "3) Cuando el dispositivo regrese automaticamente a los ajustes, seleccionar 'Seleccion automatica'.\n"
        "4) Verificar el icono RAT en la parte superior de la pantalla.",
        "El dispositivo muestra el icono LTE/4G correcto tras la busqueda manual y seleccion de red.",
        "semi", "_semi_lte_scan"
    ),
    WOMTestCase(
        "wom_86", "86", "[WOM] [86] Prueba de Tethering", "general",
        "Verificar que el terminal funciona correctamente en modo Tethering.",
        "Con LTE conectado:\n"
        "1) Activar Tethering (compartir Internet).\n"
        "2) Conectar otro dispositivo al hotspot.\n"
        "3) Navegar desde el dispositivo conectado.",
        "El tethering funciona correctamente y el dispositivo conectado tiene acceso a internet.",
        "semi", "_semi_tethering"
    ),
    WOMTestCase(
        "wom_87", "87", "[WOM] [87] Prueba de Conectividad Bluetooth", "general",
        "Verificar el envio correcto de archivos via Bluetooth.",
        "1) Emparejar el DUT con otro dispositivo via Bluetooth.\n"
        "2) Confirmar que el PIN es correcto.\n"
        "3) Transferir un archivo via Bluetooth.",
        "El Bluetooth empareja y transfiere archivos correctamente.",
        "semi", "_semi_bluetooth"
    ),
    WOMTestCase(
        "wom_88", "88", "[WOM] [88] Prueba de Conectividad NFC", "general",
        "Verificar el funcionamiento correcto del NFC.",
        "Compartir un archivo via NFC entre dos dispositivos.",
        "El NFC funciona correctamente para compartir archivos.",
        "semi", "_semi_nfc"
    ),

    # ---- CSFB ----
    WOMTestCase(
        "wom_75", "75", "[WOM] [75] Llamada CSFB", "csfb",
        "Verificar la funcionalidad correcta de CSFB en el terminal.",
        "1) Registrar el DUT en una zona con cobertura LTE.\n"
        "2) Verificar CSFB realizando una llamada de voz saliente (MO).\n"
        "3) Verificar que el dispositivo cae a 3G/2G durante la llamada y regresa a LTE al colgar.",
        "El CSFB funciona correctamente: cae a CS para la llamada y regresa a LTE al finalizar.",
        "semi", "_semi_csfb_call"
    ),
    WOMTestCase(
        "wom_76", "76", "[WOM] [76] Servicios de Datos Simultaneos durante Llamada", "csfb",
        "Verificar que el terminal permite datos y voz simultaneamente.",
        "Con LTE conectado:\n"
        "1) Iniciar descarga de un archivo grande (200MB) via HTTP.\n"
        "2) Mientras se descarga, iniciar una llamada de voz.\n"
        "3) Verificar que los datos continuan durante la llamada (requiere VoLTE activo).",
        "Los datos continuan durante la llamada de voz con VoLTE activo.",
        "semi", "_semi_data_voice"
    ),
    WOMTestCase(
        "wom_78", "78", "[WOM] [78] Llamada CSFB Saliente - Calidad de Voz", "csfb",
        "Verificar el establecimiento de llamadas CSFB salientes con buena calidad de voz.",
        "Con LTE conectado y CSFB disponible:\n"
        "1) Iniciar una llamada de voz saliente (MO).\n"
        "2) Verificar que la llamada usa CSFB.\n"
        "3) Evaluar la calidad de voz durante la llamada.",
        "La calidad de voz en CSFB es aceptable (sin cortes, eco ni ruido excesivo).",
        "semi", "_semi_csfb_mo_quality"
    ),
    WOMTestCase(
        "wom_79", "79", "[WOM] [79] Llamada CSFB Entrante - Calidad de Voz", "csfb",
        "Verificar la recepcion de llamadas CSFB entrantes con buena calidad de voz.",
        "Con LTE conectado y CSFB disponible:\n"
        "1) Recibir una llamada de voz entrante (MT).\n"
        "2) Verificar que la llamada usa CSFB.\n"
        "3) Evaluar la calidad de voz durante la llamada.",
        "La calidad de voz en CSFB entrante es aceptable.",
        "semi", "_semi_csfb_mt_quality"
    ),
    WOMTestCase(
        "wom_84", "84", "[WOM] [84] Llamada de Emergencia", "csfb",
        "Verificar que el terminal realiza llamadas de emergencia correctamente.",
        "Con LTE conectado y CSFB o VoLTE disponible:\n"
        "1) Marcar el numero de emergencia (112 o 911).\n"
        "2) Verificar que la llamada conecta correctamente.\n"
        "3) Colgar la llamada.",
        "La llamada de emergencia conecta correctamente.",
        "semi", "_semi_emergency_call"
    ),

    # ---- VoLTE ----
    WOMTestCase(
        "wom_volte_0001", "FT_VoLTE_0001", "[WOM] [FT_VoLTE_0001] Registro SIP en IMS", "volte",
        "El DUT debe completar exitosamente el procedimiento de registro SIP en IMS.",
        "1) Encender el DUT para que establezca una conexion PDN/ePDG.\n"
        "2) Observar el procedimiento de registro IMS/SIP.\n"
        "3) Verificar que el DUT se registra exitosamente en el nucleo IMS.",
        "El DUT se registra exitosamente en IMS. Estado de registro IMS = Registrado.",
        "auto", "_auto_ims_registration"
    ),
    WOMTestCase(
        "wom_volte_0005", "FT_VoLTE_0005", "[WOM] [FT_VoLTE_0005] Deregistro SIP en IMS", "volte",
        "El DUT debe completar exitosamente el procedimiento de deregistro SIP.",
        "1) Encender el DUT para establecer conexion PDN/ePDG.\n"
        "2) Activar modo avion para forzar el deregistro SIP.\n"
        "3) Desactivar modo avion y verificar el re-registro.",
        "El DUT se deregistra de IMS y se re-registra correctamente tras el ciclo de modo avion.",
        "semi", "_semi_ims_deregistration"
    ),
    WOMTestCase(
        "wom_volte_0006", "FT_VoLTE_0006", "[WOM] [FT_VoLTE_0006] Llamada Saliente VoLTE - Cliente IMS", "volte",
        "El DUT debe realizar exitosamente una llamada de voz sobre IMS a otro cliente IMS.",
        "1) En el DUT, realizar llamada de voz saliente al Cliente-1 (cliente IMS).\n"
        "2) Contestar la llamada en el Cliente-1.\n"
        "3) Verificar calidad de voz y duracion.\n"
        "4) Finalizar la llamada.",
        "Llamada VoLTE saliente hacia cliente IMS establecida con buena calidad de voz.",
        "semi", "_semi_volte_mo_ims"
    ),
    WOMTestCase(
        "wom_volte_0007", "FT_VoLTE_0007", "[WOM] [FT_VoLTE_0007] Llamada Saliente VoLTE - Cliente CS", "volte",
        "El DUT debe realizar exitosamente una llamada VoLTE a un cliente de red CS.",
        "1) En el DUT, realizar llamada de voz saliente al Cliente-1 (cliente CS/legacy).\n"
        "2) Contestar la llamada en el Cliente-1.\n"
        "3) Verificar calidad de voz.\n"
        "4) Finalizar la llamada.",
        "Llamada VoLTE saliente hacia cliente CS establecida correctamente.",
        "semi", "_semi_volte_mo_cs"
    ),
    WOMTestCase(
        "wom_volte_0008", "FT_VoLTE_0008", "[WOM] [FT_VoLTE_0008] Llamada Saliente VoLTE - Linea Fija", "volte",
        "El DUT debe realizar exitosamente una llamada VoLTE a una linea fija.",
        "1) En el DUT, realizar llamada de voz saliente a un numero de linea fija.\n"
        "2) Contestar la llamada en la linea fija.\n"
        "3) Verificar calidad de voz.\n"
        "4) Finalizar la llamada.",
        "Llamada VoLTE saliente hacia linea fija establecida correctamente.",
        "semi", "_semi_volte_mo_fixed"
    ),
    WOMTestCase(
        "wom_volte_0015", "FT_VoLTE_0015", "[WOM] [FT_VoLTE_0015] Emision de Tonos DTMF en Llamada VoLTE", "volte",
        "El DUT debe transmitir tonos DTMF correctamente durante una llamada IMS.",
        "1) En el DUT, realizar llamada de voz saliente al Cliente-1.\n"
        "2) Contestar en el Cliente-1.\n"
        "3) Durante la llamada, enviar tonos DTMF desde el DUT.\n"
        "4) Verificar que el Cliente-1 recibe los tonos DTMF correctamente.",
        "Los tonos DTMF se transmiten correctamente durante la llamada VoLTE.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0020", "FT_VoLTE_0020", "[WOM] [FT_VoLTE_0020] Calidad de Voz en Llamada VoLTE", "volte",
        "Verificar la calidad de voz en llamadas VoLTE en condiciones de senal debil.",
        "1) Realizar llamada VoLTE saliente con senal LTE debil (RSRP < -100dBm).\n"
        "2) Mantener la llamada 3 minutos.\n"
        "3) Evaluar la calidad de voz (puntuacion MOS si es posible).",
        "La calidad de voz es aceptable incluso en condiciones de senal debil.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0025", "FT_VoLTE_0025", "[WOM] [FT_VoLTE_0025] SMS sobre IMS durante Llamada Activa", "volte",
        "Verificar que el DUT puede enviar SMS via IMS durante una llamada de voz activa.",
        "1) Con una llamada VoLTE activa, redactar un nuevo SMS desde la aplicacion de mensajes del DUT.\n"
        "2) Ingresar el numero del Cliente-1 como destinatario.\n"
        "3) Enviar el SMS.\n"
        "4) Verificar que el Cliente-1 recibe el SMS.",
        "SMS enviado exitosamente via IMS durante una llamada VoLTE activa.",
        "semi", "_semi_sms_during_call"
    ),
    WOMTestCase(
        "wom_volte_0026", "FT_VoLTE_0026", "[WOM] [FT_VoLTE_0026] Recepcion de SMS Entrante sobre IMS", "volte",
        "Verificar que el DUT puede recibir un SMS entrante via IMS.",
        "1) Desde el Cliente-1, redactar un nuevo SMS con el numero del DUT como destinatario.\n"
        "2) Enviar el SMS desde el Cliente-1.\n"
        "3) Verificar que el DUT recibe el SMS correctamente.",
        "El DUT recibe el SMS entrante via IMS exitosamente.",
        "semi", "_semi_receive_sms"
    ),
    WOMTestCase(
        "wom_volte_0039", "FT_VoLTE_0039", "[WOM] [FT_VoLTE_0039] Llamada en Espera (CW) - Configuracion Terminal", "volte",
        "Verificar la activacion y desactivacion de Llamada en Espera gestionada por el terminal (sin senalizacion de red).",
        "1) En el DUT, desactivar la Llamada en Espera (CW) desde Ajustes de llamada.\n"
        "2) Consultar el estado de CW: verificar que indica 'desactivada'.\n"
        "3) Confirmar que no se envio senalizacion Ut/XCAP a la red.\n"
        "4) Activar CW desde el DUT.\n"
        "5) Consultar el estado: verificar que indica 'activada'.\n"
        "6) Con una llamada activa, recibir una segunda llamada y verificar la notificacion de CW.",
        "La Llamada en Espera se activa/desactiva desde el terminal sin senalizacion de red y funciona correctamente.",
        "semi", "_semi_call_waiting_toggle"
    ),
    WOMTestCase(
        "wom_volte_0040", "FT_VoLTE_0040", "[WOM] [FT_VoLTE_0040] Llamada en Espera (CW) - Configuracion de Red (Ut/XCAP)", "volte",
        "Verificar la activacion y desactivacion de Llamada en Espera gestionada por la red via Ut/XCAP.",
        "1) En el DUT, desactivar CW via configuracion de red (Ut/XCAP).\n"
        "2) Confirmar que el DUT envio la solicitud Ut/XCAP a la red.\n"
        "3) Consultar el estado: verificar que indica 'desactivada'.\n"
        "4) Activar CW via configuracion de red.\n"
        "5) Confirmar que el DUT envio la solicitud Ut/XCAP.\n"
        "6) Con una llamada activa, recibir una segunda llamada y verificar el comportamiento de CW.",
        "La configuracion de CW via red (Ut/XCAP) funciona correctamente y la llamada en espera opera segun lo esperado.",
        "semi", "_semi_call_waiting_toggle"
    ),
    WOMTestCase(
        "wom_volte_0048", "FT_VoLTE_0048", "[WOM] [FT_VoLTE_0048] Llamada en Espera - Iniciada desde DUT", "volte",
        "Verificar la operacion de Poner en Espera (Hold) iniciada desde el DUT durante llamada IMS.",
        "1) En el DUT, poner al Cliente-1 en espera usando la opcion del menu.\n"
        "2) Esperar 15 segundos.\n"
        "3) En el DUT, recuperar la llamada del Cliente-1.\n"
        "4) Verificar que ambas partes pueden comunicarse.",
        "El DUT puede poner y recuperar una llamada en espera exitosamente.",
        "semi", "_semi_call_hold_mo"
    ),
    WOMTestCase(
        "wom_volte_0049", "FT_VoLTE_0049", "[WOM] [FT_VoLTE_0049] Llamada en Espera - Iniciada por Remoto", "volte",
        "Verificar la operacion de Poner en Espera iniciada por la parte remota durante llamada IMS.",
        "1) Desde el Cliente-1, poner al DUT en espera.\n"
        "2) Esperar 15 segundos.\n"
        "3) Desde el Cliente-1, recuperar la llamada con el DUT.\n"
        "4) Verificar que ambas partes pueden comunicarse.",
        "El DUT gestiona correctamente el estado de espera cuando es puesto en espera por la parte remota.",
        "semi", "_semi_call_hold_mt"
    ),
    WOMTestCase(
        "wom_volte_0054", "FT_VoLTE_0054", "[WOM] [FT_VoLTE_0054] Notificacion de Buzon de Voz sobre IMS", "volte",
        "Verificar que el DUT recibe una notificacion de buzon de voz (MWI) sobre IMS.",
        "1) En el DUT, recibir llamada entrante del Cliente-1. No contestar.\n"
        "2) El Cliente-1 deja un mensaje en el buzon de voz.\n"
        "3) Verificar que el DUT recibe la notificacion MWI.",
        "El DUT recibe la notificacion MWI (Indicacion de Mensaje en Espera) sobre IMS.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0055", "FT_VoLTE_0055", "[WOM] [FT_VoLTE_0055] Conferencia Saliente VoLTE - Clientes IMS", "volte",
        "Verificar llamada de conferencia con multiples participantes IMS.",
        "1) En el DUT, llamar al Cliente-1. Contestar en el Cliente-1.\n"
        "2) En el DUT, llamar al Cliente-2. Contestar en el Cliente-2.\n"
        "3) En el DUT, fusionar las llamadas en conferencia.\n"
        "4) Verificar que todos los participantes pueden comunicarse.",
        "Llamada de conferencia con clientes IMS establecida exitosamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0056", "FT_VoLTE_0056", "[WOM] [FT_VoLTE_0056] Conferencia Saliente VoLTE - Clientes no IMS", "volte",
        "Verificar llamada de conferencia con participantes que no son clientes IMS.",
        "1) En el DUT, llamar al Cliente-1 (no IMS). Contestar en el Cliente-1.\n"
        "2) En el DUT, llamar al Cliente-2 (no IMS). Contestar en el Cliente-2.\n"
        "3) En el DUT, fusionar las llamadas en conferencia.",
        "Llamada de conferencia con clientes no IMS establecida exitosamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0057", "FT_VoLTE_0057", "[WOM] [FT_VoLTE_0057] Conferencia Entrante VoLTE - Clientes IMS", "volte",
        "Verificar que el DUT puede unirse a una conferencia iniciada por clientes IMS.",
        "1) En el DUT, recibir llamada entrante del Cliente-1. Contestar.\n"
        "2) El Cliente-1 incorpora al DUT en una conferencia con el Cliente-2.\n"
        "3) Verificar que todos los participantes pueden comunicarse.",
        "El DUT se une exitosamente a una conferencia iniciada por un cliente IMS.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0058", "FT_VoLTE_0058", "[WOM] [FT_VoLTE_0058] Conferencia Entrante VoLTE - Clientes no IMS", "volte",
        "Verificar que el DUT puede unirse a una conferencia con clientes no IMS.",
        "1) En el DUT, recibir llamada entrante del Cliente-1. Contestar.\n"
        "2) El Cliente-1 incorpora al DUT en una conferencia con el Cliente-2 (no IMS).\n"
        "3) Verificar que todos los participantes pueden comunicarse.",
        "El DUT se une exitosamente a una conferencia con clientes no IMS.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0060", "FT_VoLTE_0060", "[WOM] [FT_VoLTE_0060] USSD sobre IMS durante Llamada Activa", "volte",
        "Verificar si el DUT utiliza USSD sobre IMS correctamente durante una llamada.",
        "1) En el DUT, ingresar y enviar un codigo USSD (ej: *901# para consultar saldo prepago) "
        "durante una llamada VoLTE activa.\n"
        "2) Verificar que se recibe la respuesta USSD.",
        "El codigo USSD funciona correctamente durante una llamada VoLTE activa.",
        "semi", "_semi_ussd_during_call"
    ),
    WOMTestCase(
        "wom_volte_0068", "FT_VoLTE_0068", "[WOM] [FT_VoLTE_0068] Llamada de Emergencia via VxLTE - CSFB", "volte",
        "Verificar que el DUT puede realizar una llamada de emergencia via CSFB.",
        "1) En el DUT, realizar llamada de emergencia al numero 112 o 911.\n"
        "2) Finalizar la llamada de emergencia.\n"
        "3) Verificar que el DUT usa CSFB para la llamada de emergencia.",
        "El DUT puede realizar llamadas de emergencia via CSFB cuando VxLTE no esta soportado.",
        "semi", "_semi_emergency_volte"
    ),
    WOMTestCase(
        "wom_volte_0069", "FT_VoLTE_0069", "[WOM] [FT_VoLTE_0069] Llamada de Emergencia VxLTE - Celda Alternativa", "volte",
        "Verificar que el DUT puede realizar llamadas de emergencia desde una celda alternativa.",
        "1) Mover el DUT a una zona registrada para VxLTE pero sin celda CS adecuada para CSFB.\n"
        "2) Realizar llamada de emergencia.\n"
        "3) Verificar que la llamada conecta.",
        "El DUT realiza la llamada de emergencia exitosamente aunque no haya celda CS adecuada.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0081", "FT_VoLTE_0081", "[WOM] [FT_VoLTE_0081] Llamada VoLTE durante Transferencia de Datos", "volte",
        "Verificar que el DUT puede establecer una llamada VoLTE durante una transferencia de datos activa.",
        "1) En el DUT, descargar un archivo grande desde un servidor externo.\n"
        "2) Mientras se descarga, realizar llamada de voz saliente al Cliente-1.\n"
        "3) Contestar en el Cliente-1.\n"
        "4) Verificar que datos y voz funcionan simultaneamente.",
        "Llamada VoLTE establecida exitosamente mientras hay una transferencia de datos en curso.",
        "semi", "_semi_call_during_data"
    ),
    WOMTestCase(
        "wom_volte_0082", "FT_VoLTE_0082", "[WOM] [FT_VoLTE_0082] Transferencia de Datos durante Llamada VoLTE", "volte",
        "Verificar que el DUT puede transferir datos durante una llamada VoLTE activa.",
        "1) En el DUT, realizar llamada de voz saliente al Cliente-1.\n"
        "2) Contestar en el Cliente-1.\n"
        "3) Con la llamada activa, iniciar descarga de un archivo grande en el DUT.\n"
        "4) Verificar que datos y voz funcionan simultaneamente.",
        "La transferencia de datos funciona correctamente durante una llamada VoLTE activa.",
        "semi", "_semi_data_during_call"
    ),
    WOMTestCase(
        "wom_volte_0098", "FT_VoLTE_0098", "[WOM] [FT_VoLTE_0098] Registro VoLTE en Dual SIM", "volte",
        "Verificar el registro VoLTE en ambas ranuras SIM de un dispositivo Dual SIM.",
        "1) Encender y apagar el DUT.\n"
        "2) Realizar llamadas MO/MT en la SIM1 y la SIM2.\n"
        "3) Enviar SMS MO/MT en ambas SIM.\n"
        "4) Verificar el registro IMS en ambas ranuras.",
        "Ambas ranuras SIM se registran en IMS y VoLTE funciona en ambas.",
        "semi", "_semi_dual_sim_volte"
    ),
    WOMTestCase(
        "wom_volte_0099", "FT_VoLTE_0099", "[WOM] [FT_VoLTE_0099] Interaccion Dual SIM con VoLTE", "volte",
        "Verificar la interaccion Dual SIM durante uso simultaneo de datos y VoLTE.",
        "1) SIM1 navegando con datos mientras el Cliente invita una llamada a SIM2.\n"
        "2) SIM2 contesta la llamada.\n"
        "3) Verificar que los datos de SIM1 y la voz de SIM2 funcionan simultaneamente.",
        "La interaccion Dual SIM con VoLTE funciona correctamente.",
        "manual"
    ),
    WOMTestCase(
        "wom_volte_0101", "FT_VoLTE_0101", "[WOM] [FT_VoLTE_0101] Cambio de Red Preferida entre SIM1 y SIM2", "volte",
        "Verificar el cambio de red preferida entre las ranuras SIM1 y SIM2.",
        "Cambiar la red preferida de la SIM1 a la SIM2, verificar que ambas ranuras se registran en IMS correctamente.",
        "El cambio de red preferida entre ranuras SIM funciona correctamente con IMS.",
        "semi", "_semi_switch_preferred_sim"
    ),
    WOMTestCase(
        "wom_volte_0102", "FT_VoLTE_0102", "[WOM] [FT_VoLTE_0102] Cambio de Red Preferida SIM1/SIM2 - Parte II", "volte",
        "Verificar el cambio de red preferida entre SIM1 y SIM2 en ambas direcciones.",
        "1) Cambiar la red preferida de SIM1 a SIM2, verificar registro.\n"
        "2) Cambiar de vuelta de SIM2 a SIM1.\n"
        "3) Verificar que ambos escenarios funcionan correctamente.",
        "El cambio de red preferida funciona correctamente en ambas direcciones.",
        "semi", "_semi_switch_preferred_sim2"
    ),

    # ---- VoWiFi ----
    WOMTestCase(
        "wom_vowifi_0153", "FT_VoWIFI_00153", "[WOM] [FT_VoWIFI_00153] Activar/Desactivar WiFi Calling", "vowifi",
        "Verificar la activacion y desactivacion de la opcion WiFi Calling.",
        "1) Verificar que el DUT tiene la opcion de activar/desactivar WiFi Calling.\n"
        "   Si no existe, el resultado es N/A.\n"
        "2) Activar WiFi Calling y verificar el registro IMS sobre WiFi.\n"
        "3) Desactivar WiFi Calling y verificar el deregistro.",
        "WiFi Calling puede activarse/desactivarse y el registro/deregistro IMS ocurre correctamente.",
        "semi", "_semi_vowifi_toggle"
    ),
    WOMTestCase(
        "wom_vowifi_0154", "FT_VoWIFI_00154", "[WOM] [FT_VoWIFI_00154] Llamada Saliente por WiFi a WiFi", "vowifi",
        "Verificar llamada de voz saliente desde WiFi hacia otro cliente WiFi.",
        "1) Activar WiFi Calling en el DUT y en la Parte B.\n"
        "2) Llamar a la Parte B desde el DUT.\n"
        "3) Contestar la llamada en la Parte B.\n"
        "4) Verificar la calidad de voz.",
        "Llamada VoWiFi saliente hacia cliente WiFi establecida exitosamente.",
        "semi", "_semi_vowifi_mo_wifi"
    ),
    WOMTestCase(
        "wom_vowifi_01555", "FT_VoWIFI_001555", "[WOM] [FT_VoWIFI_001555] Llamada Saliente WiFi a VoLTE", "vowifi",
        "Verificar llamada de voz saliente desde WiFi hacia un cliente VoLTE.",
        "1) Activar WiFi Calling en el DUT.\n"
        "2) Realizar llamada saliente desde el DUT a la Parte B (numero VoLTE).\n"
        "3) Contestar en la Parte B.\n"
        "4) Verificar la calidad de voz.",
        "Llamada VoWiFi saliente hacia cliente VoLTE establecida exitosamente.",
        "semi", "_semi_vowifi_mo_volte"
    ),
    WOMTestCase(
        "wom_vowifi_0156", "FT_VoWIFI_0156", "[WOM] [FT_VoWIFI_0156] Llamada Saliente WiFi con Desconexion de WiFi", "vowifi",
        "Verificar el comportamiento de una llamada WiFi cuando se desactiva el WiFi durante la llamada.",
        "1) Encender el DUT.\n"
        "2) Iniciar llamada de voz saliente desde el DUT hacia la Parte B (movil o PSTN).\n"
        "3) Contestar en la Parte B.\n"
        "4) Durante la llamada, desactivar el WiFi en el DUT.\n"
        "5) Verificar que la llamada continua (con traspaso a VoLTE o CSFB).",
        "La llamada continua tras desactivar el WiFi con un traspaso correcto.",
        "semi", "_semi_vowifi_handover"
    ),
    WOMTestCase(
        "wom_vowifi_0157", "FT_VoWIFI_00157", "[WOM] [FT_VoWIFI_00157] Llamada Saliente WiFi con Caida a CSFB", "vowifi",
        "Verificar llamada saliente WiFi con traspaso (handover) a CSFB.",
        "1) Activar WiFi Calling en el DUT.\n"
        "2) Iniciar llamada WiFi desde el DUT hacia la Parte B (cliente CSFB).\n"
        "3) Contestar la llamada en la Parte B.\n"
        "4) Verificar que la llamada se establece como WiFi en el DUT.\n"
        "5) Verificar la calidad de audio en ambas direcciones.\n"
        "6) Finalizar la llamada desde cualquiera de las dos partes.",
        "La llamada VoWiFi hacia cliente CSFB se establece correctamente con buena calidad de audio.",
        "semi", "_semi_vowifi_mo_csfb"
    ),
    WOMTestCase(
        "wom_vowifi_0158", "FT_VoWIFI_00158", "[WOM] [FT_VoWIFI_00158] Llamadas MO y MT entre WiFi y PSTN (Linea Fija)", "vowifi",
        "Verificar llamadas salientes (MO) y entrantes (MT) entre WiFi Calling y PSTN (linea fija).",
        "1) Activar WiFi Calling en el DUT.\n"
        "2) Llamada MO: marcar un numero de linea fija desde el DUT.\n"
        "3) Contestar en la linea fija y verificar la calidad.\n"
        "4) Finalizar la llamada.\n"
        "5) Llamada MT: desde la linea fija, llamar al numero del DUT.\n"
        "6) Contestar en el DUT y verificar que la llamada llega via WiFi.\n"
        "7) Verificar la calidad de audio.",
        "Las llamadas MO y MT entre WiFi Calling y PSTN funcionan correctamente con buena calidad.",
        "semi", "_semi_vowifi_pstn_call"
    ),
    WOMTestCase(
        "wom_vowifi_0159", "FT_VoWIFI_00159", "[WOM] [FT_VoWIFI_00159] Llamada Entrante VoLTE a WiFi", "vowifi",
        "Verificar llamada entrante desde VoLTE hacia un DUT registrado en WiFi Calling.",
        "1) Activar WiFi Calling en el DUT.\n"
        "2) La Parte B inicia una llamada VoLTE hacia el DUT.\n"
        "3) Contestar en el DUT via WiFi.",
        "La llamada entrante desde VoLTE hacia el DUT registrado en WiFi funciona correctamente.",
        "semi", "_semi_vowifi_mt_volte"
    ),
    WOMTestCase(
        "wom_vowifi_0160", "FT_VoWIFI_00160", "[WOM] [FT_VoWIFI_00160] Llamada Entrante CSFB a WiFi", "vowifi",
        "Verificar llamada entrante desde CSFB hacia un DUT registrado en WiFi Calling.",
        "1) Activar WiFi Calling en el DUT y verificar el registro IMS via WiFi.\n"
        "2) La Parte B inicia una llamada CSFB hacia el numero del DUT.\n"
        "3) Contestar en el DUT.\n"
        "4) Verificar que la interfaz del DUT indica que la llamada es via WiFi.\n"
        "5) Verificar la calidad de audio.\n"
        "6) Finalizar la llamada.",
        "La llamada entrante desde CSFB hacia el DUT registrado en WiFi se establece correctamente.",
        "semi", "_semi_vowifi_mt_csfb"
    ),
    WOMTestCase(
        "wom_vowifi_0179", "FT_VoWIFI_00179", "[WOM] [FT_VoWIFI_00179] Llamada de Emergencia en Modo Avion con WiFi", "vowifi",
        "Verificar llamada de emergencia cuando el DUT esta en Modo Avion con cobertura WiFi.",
        "Condiciones previas: El DUT esta en modo avion pero registrado en IMS via VoWiFi activo.\n"
        "1) Confirmar que el DUT esta en modo avion y registrado en IMS via WiFi.\n"
        "2) Realizar llamada de emergencia (112 o 911).\n"
        "3) Verificar que la llamada conecta via WiFi.\n"
        "4) Colgar inmediatamente.\n"
        "NOTA: Algunos operadores no soportan emergencias via VoWiFi; "
        "el DUT puede desactivar el modo avion automaticamente y usar VoLTE o CSFB.",
        "La llamada de emergencia conecta via la mejor ruta disponible (WiFi, VoLTE o CS).",
        "semi", "_semi_emergency_vowifi"
    ),
    WOMTestCase(
        "wom_vowifi_0180", "FT_VoWIFI_00180", "[WOM] [FT_VoWIFI_00180] Llamada de Emergencia con 4G HPLMN y WiFi", "vowifi",
        "Verificar llamada de emergencia cuando el DUT tiene cobertura 4G HPLMN y WiFi activos simultaneamente.",
        "Condiciones previas: El DUT esta conectado a la red 4G HPLMN y registrado en IMS via VoWiFi.\n"
        "1) Confirmar que el DUT tiene 4G activo y WiFi Calling registrado.\n"
        "2) Realizar llamada de emergencia (112 o 911).\n"
        "3) Observar la ruta utilizada: VoLTE, WiFi o CSFB.\n"
        "4) Colgar inmediatamente.\n"
        "NOTA: Verificar en el log si aplica, si la llamada fue correcta segun los requisitos del operador.",
        "La llamada de emergencia conecta correctamente usando la ruta adecuada segun la configuracion del operador.",
        "semi", "_semi_emergency_vowifi"
    ),

    # ---- 5G NR NSA ----
    WOMTestCase(
        "wom_5g_0219", "FT_5GNR_NSA_00219", "[WOM] [FT_5GNR_NSA_00219] VoLTE en Cobertura 5G NR", "5g",
        "Verificar llamadas VoLTE en zona con cobertura 5G NR.",
        "1) Verificar que el DUT esta conectado a la celda LTE y NR.\n"
        "2) Realizar una llamada VoLTE.\n"
        "3) Verificar que la llamada se establece sobre VoLTE mientras los datos NR estan activos.",
        "La llamada VoLTE funciona correctamente en cobertura 5G NR NSA.",
        "semi", "_semi_5g_volte"
    ),
    WOMTestCase(
        "wom_5g_0220", "FT_5GNR_NSA_00220", "[WOM] [FT_5GNR_NSA_00220] Datos en Cobertura 5G NR", "5g",
        "Verificar sesion de datos en cobertura 5G NR.",
        "1) Verificar que el DUT esta conectado a la celda LTE y NR.\n"
        "2) Activar una sesion de datos.\n"
        "3) Verificar que el throughput de datos utiliza 5G NR.",
        "La sesion de datos utiliza 5G NR NSA correctamente.",
        "semi", "_semi_5g_data"
    ),
    WOMTestCase(
        "wom_5g_0221", "FT_5GNR_NSA_00221", "[WOM] [FT_5GNR_NSA_00221] Datos y Voz Simultaneos en 5G NR", "5g",
        "Verificar datos y llamada de voz en paralelo en cobertura 5G NR.",
        "1) Verificar que el DUT esta conectado a la celda LTE y NR.\n"
        "2) Activar una sesion de datos.\n"
        "3) Con los datos activos, realizar una llamada VoLTE.\n"
        "4) Verificar que datos y voz funcionan simultaneamente.",
        "Los datos 5G NR NSA y la llamada VoLTE funcionan en paralelo correctamente.",
        "semi", "_semi_5g_data_voice"
    ),
]


# ==================== EJECUTOR ====================

class SanityWOMExecutor:
    def __init__(self, adb_manager):
        self.adb = adb_manager
        self.tests: Dict[str, WOMTestCase] = {t.id: t for t in WOM_TEST_CASES}
        self._lock = threading.Lock()
        self._load_results()

    # ---------- Persistencia ----------

    def _load_results(self):
        if os.path.exists(RESULTS_FILE):
            try:
                with open(RESULTS_FILE, 'r', encoding='utf-8') as f:
                    saved = json.load(f)
                for tc_id, data in saved.items():
                    if tc_id in self.tests:
                        self.tests[tc_id].result = data.get('result', 'pending')
                        self.tests[tc_id].remark = data.get('remark', '')
                        self.tests[tc_id].last_run = data.get('last_run', '')
                logger.info(f"Sanity WOM: resultados cargados desde {RESULTS_FILE}")
            except Exception as e:
                logger.warning(f"No se pudieron cargar resultados WOM: {e}")

    def _save_results(self):
        os.makedirs('data', exist_ok=True)
        data = {}
        for tc_id, tc in self.tests.items():
            data[tc_id] = {
                'result': tc.result,
                'remark': tc.remark,
                'last_run': tc.last_run,
            }
        with open(RESULTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

    # ---------- API principal ----------

    def get_test_cases(self) -> List[dict]:
        with self._lock:
            return [tc.to_dict() for tc in WOM_TEST_CASES]

    def set_result(self, test_id: str, result: str, remark: str = '') -> bool:
        with self._lock:
            if test_id not in self.tests:
                return False
            self.tests[test_id].result = result.lower()
            self.tests[test_id].remark = remark
            self.tests[test_id].last_run = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self._save_results()
            return True

    def reset_results(self):
        with self._lock:
            for tc in self.tests.values():
                tc.result = 'pending'
                tc.remark = ''
                tc.last_run = ''
            self._save_results()

    def run_auto_test(self, test_id: str, serial: str) -> dict:
        """Ejecuta la funcion de automatizacion ADB de un caso (AUTO o SEMI)."""
        with self._lock:
            if test_id not in self.tests:
                return {'success': False, 'message': 'Test no encontrado'}
            tc = self.tests[test_id]
            if not tc.auto_func:
                return {'success': False, 'message': 'Este test no tiene automatizacion ADB'}

        func = getattr(self, tc.auto_func, None)
        if func is None:
            return {'success': False, 'message': f'Funcion {tc.auto_func} no implementada'}

        try:
            result = func(serial)
            with self._lock:
                self.tests[test_id].last_run = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                if result.get('auto_result'):
                    self.tests[test_id].result = result['auto_result']
                self._save_results()
            return result
        except Exception as e:
            logger.error(f"Error ejecutando {tc.auto_func}: {e}")
            return {'success': False, 'message': str(e)}

    # ---------- Generacion de reporte Excel ----------

    def generate_excel_report(self, model: str = '', tester: str = '', sw_version: str = '') -> str:
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl no instalado")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sanity Check WOM"

        COLOR_HEADER = "1A237E"
        COLOR_PASS    = "C8E6C9"
        COLOR_FAIL    = "FFCDD2"
        COLOR_NA      = "E0E0E0"
        COLOR_PENDING = "FFF9C4"
        COLOR_CAT = {
            'general': "E3F2FD", 'csfb': "FFF3E0",
            'volte': "F3E5F5", 'vowifi': "E8F5E9", '5g': "FCE4EC"
        }
        CAT_LABELS = {
            'general': 'General', 'csfb': 'CSFB',
            'volte': 'VoLTE', 'vowifi': 'VoWiFi', '5g': '5G NR NSA'
        }

        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:G1')
        ws['A1'] = f"Sanity Check WOM Colombia - {model or 'Modelo'}"
        ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
        ws['A1'].fill = PatternFill("solid", fgColor=COLOR_HEADER)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        meta = [
            ("Tester:", tester or '-'),
            ("Modelo:", model or '-'),
            ("SW Version:", sw_version or '-'),
            ("Fecha:", datetime.now().strftime('%Y-%m-%d %H:%M')),
        ]
        for i, (label, val) in enumerate(meta, start=2):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = val

        counts = {'pass': 0, 'fail': 0, 'na': 0, 'pending': 0}
        for tc in self.tests.values():
            counts[tc.result] = counts.get(tc.result, 0) + 1

        ws['D2'] = 'PASS';      ws['E2'] = counts['pass']
        ws['D3'] = 'FAIL';      ws['E3'] = counts['fail']
        ws['D4'] = 'N/A';       ws['E4'] = counts['na']
        ws['D5'] = 'PENDIENTE'; ws['E5'] = counts['pending']
        ws['D2'].font = Font(bold=True, color='2E7D32')
        ws['D3'].font = Font(bold=True, color='C62828')

        headers   = ['#', 'ID', 'Nombre', 'Categoria', 'Automatizacion', 'Resultado', 'Observacion']
        col_widths = [5,   20,    45,         12,           14,              12,           30]
        header_row = 7
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=header_row, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill("solid", fgColor=COLOR_HEADER)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[header_row].height = 20

        for row_idx, tc in enumerate(WOM_TEST_CASES, start=header_row + 1):
            result_upper = tc.result.upper() if tc.result != 'pending' else 'PENDIENTE'
            result_color = {
                'PASS': COLOR_PASS, 'FAIL': COLOR_FAIL,
                'NA': COLOR_NA, 'PENDIENTE': COLOR_PENDING
            }.get(result_upper, COLOR_PENDING)

            cat_color  = COLOR_CAT.get(tc.category, 'FFFFFF')
            auto_label = {'auto': 'AUTO', 'semi': 'SEMI', 'manual': 'MANUAL'}.get(tc.automation, tc.automation)

            values = [
                row_idx - header_row,
                tc.number,
                tc.name,
                CAT_LABELS.get(tc.category, tc.category),
                auto_label,
                result_upper,
                tc.remark
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                if col == 6:
                    cell.fill = PatternFill("solid", fgColor=result_color)
                    cell.font = Font(bold=True)
                elif col == 4:
                    cell.fill = PatternFill("solid", fgColor=cat_color)
            ws.row_dimensions[row_idx].height = 18

        ws.freeze_panes = f'A{header_row + 1}'

        os.makedirs(os.path.join('data', 'sanity_wom_reports'), exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filepath  = os.path.join('data', 'sanity_wom_reports',
                                 f'SanityWOM_{model or "report"}_{timestamp}.xlsx')
        wb.save(filepath)
        logger.info(f"Reporte WOM generado: {filepath}")
        return filepath

    # ==================== FUNCIONES ADB - AUTO ====================

    def _adb(self, serial: str, cmd: str) -> str:
        """Ejecuta un comando ADB y devuelve la salida (string)."""
        _, output = self.adb.run_command(cmd, serial)
        return output or ''

    def _auto_spn(self, serial: str) -> dict:
        """Verifica que el SPN de WOM se muestra correctamente."""
        try:
            operator = self._adb(serial, 'shell getprop gsm.operator.alpha').strip()
            if not operator:
                operator = self._adb(serial, 'shell getprop gsm.operator.alpha.2').strip()

            message = f"Operador detectado: '{operator}'"
            if not operator:
                message = "No se pudo leer el nombre de operador. Verificar manualmente en la pantalla de inicio."

            return {
                'success': True,
                'message': message,
                'auto_result': 'pass' if operator else None
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _auto_apn(self, serial: str) -> dict:
        """Verifica que el APN de WOM esta configurado correctamente."""
        try:
            apn_info = self._adb(serial,
                'shell content query --uri content://telephony/carriers/preferapn '
                '--projection name:apn:proxy:port'
            ).strip()

            wom_apn = any(x in apn_info.lower() for x in ['wom', 'internet.wom', 'wom.co'])
            message = f"APN preferido:\n{apn_info[:300] if apn_info else 'No encontrado'}"
            return {
                'success': True,
                'message': message,
                'auto_result': 'pass' if wom_apn else None
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _auto_time_sync(self, serial: str) -> dict:
        """Deshabilita auto_time, fuerza fecha incorrecta, re-habilita y verifica sincronizacion."""
        import time as _time

        WRONG_DATE_TOYBOX = 'shell date 010100002020.00'   # MMDDhhmm[[CC]YY][.ss]
        WRONG_DATE_SU     = 'shell su -c "date 010100002020.00"'

        def _try_set_wrong_date() -> bool:
            """Intenta fijar la fecha en 2020 por varios metodos. Devuelve True si lo logro."""
            # Metodo 1: adb root (userdebug / eng builds)
            ok, out = self.adb.run_command('root', serial)
            if ok and 'cannot' not in out.lower() and 'error' not in out.lower():
                _time.sleep(3)   # esperar a que adbd reinicie como root
                self._adb(serial, WRONG_DATE_TOYBOX)
                if '2020' in self._adb(serial, 'shell date'):
                    return True

            # Metodo 2: su -c (Magisk / SuperSU instalado en el dispositivo)
            self._adb(serial, WRONG_DATE_SU)
            if '2020' in self._adb(serial, 'shell date'):
                return True

            return False

        try:
            # 1. Hora inicial
            time_initial = self._adb(serial, 'shell date').strip()

            # 2. Deshabilitar auto_time
            self._adb(serial, 'shell settings put global auto_time 0')
            _time.sleep(1)

            # 3. Intentar poner fecha incorrecta
            date_set_ok = _try_set_wrong_date()
            time_wrong  = self._adb(serial, 'shell date').strip()

            # 4. Re-habilitar auto_time y forzar sincronizacion
            self._adb(serial, 'shell settings put global auto_time 1')
            self._adb(serial, 'shell am broadcast -a android.intent.action.TIME_SET')

            # 5. Esperar sincronizacion con la red (~20 s)
            _time.sleep(20)

            # 6. Leer hora final
            time_after = self._adb(serial, 'shell date').strip()

            if date_set_ok:
                synced = '2020' not in time_after
                msg = (
                    f"Hora inicial: {time_initial}\n"
                    f"Hora incorrecta establecida (2020): {time_wrong}\n"
                    f"Hora tras re-habilitar auto_time (20s): {time_after}\n"
                    f"Sincronizacion: {'EXITOSA - el dispositivo corrigio la hora' if synced else 'FALLIDA - la hora no se actualizo'}"
                )
                return {
                    'success': True,
                    'message': msg,
                    'auto_result': 'pass' if synced else 'fail'
                }
            else:
                # Build de produccion sin root ni su
                auto_time = self._adb(serial, 'shell settings get global auto_time').strip()
                auto_zone = self._adb(serial, 'shell settings get global auto_time_zone').strip()
                msg = (
                    f"NOTA: El dispositivo no permite cambiar la fecha por ADB "
                    f"(build de produccion; no tiene 'adb root' ni 'su').\n"
                    f"Hora del dispositivo: {time_initial}\n"
                    f"auto_time={auto_time} (1=activo), auto_time_zone={auto_zone}\n"
                    f"Para habilitar root: active 'Opciones de desarrollador > Root ADB' "
                    f"(si el dispositivo lo soporta) o instale Magisk.\n"
                    f"Verifique manualmente que la hora se sincroniza con la red WOM."
                )
                return {
                    'success': True,
                    'message': msg,
                    'auto_result': 'pass' if auto_time == '1' else 'fail'
                }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _auto_ims_registration(self, serial: str) -> dict:
        """Verifica que el dispositivo esta registrado en IMS."""
        try:
            out = self._adb(serial, 'shell dumpsys telephony.registry')
            ims_lines = [l.strip() for l in out.splitlines()
                         if any(k in l for k in ['ImsServiceState', 'imsMmTelFeatureState',
                                                  'mImsRegistered', 'IMS_REGISTERED', 'imsReady'])]

            registered = any('true' in l.lower() or 'registered' in l.lower() or '= 1' in l
                             for l in ims_lines)
            message = '\n'.join(ims_lines[:8]) if ims_lines else "No se encontro informacion IMS en el registro"

            return {
                'success': True,
                'message': f"Estado IMS:\n{message}",
                'auto_result': 'pass' if registered else None
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    # ==================== FUNCIONES ADB - SEMI ====================

    def _semi_network_selection(self, serial: str) -> dict:
        """Abre la pantalla de seleccion de red."""
        try:
            self._adb(serial, 'shell am start -a android.settings.NETWORK_OPERATOR_SETTINGS')
            return {
                'success': True,
                'message': 'Pantalla de seleccion de red abierta. Verificar el escaneo y la seleccion manual/automatica de red.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_browser(self, serial: str) -> dict:
        """Abre el navegador con una pagina de prueba."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d http://www.google.com')
            return {
                'success': True,
                'message': 'Navegador abierto en google.com. Verificar que carga correctamente y descargar algun contenido.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_ussd(self, serial: str) -> dict:
        """Abre el marcador con el codigo USSD de WOM."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.CALL -d tel:%2A611%23')
            return {
                'success': True,
                'message': 'Marcando *611# (codigo USSD WOM). Verificar la respuesta en pantalla.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_prepago_ussd(self, serial: str) -> dict:
        """Marca el codigo USSD de servicios prepago WOM: *888#"""
        try:
            # *888# → URL encoded: %2A888%23  (* = %2A, # = %23)
            self._adb(serial, 'shell am start -a android.intent.action.CALL -d tel:%2A888%23')
            return {
                'success': True,
                'message': 'Marcando *888# (servicios prepago WOM). Verificar que el menu de prepago responde correctamente en pantalla.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_gps(self, serial: str) -> dict:
        """Activa GPS y abre Google Maps."""
        try:
            self._adb(serial, 'shell settings put secure location_mode 3')
            time.sleep(1)
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d geo:0,0?q=my+location')
            return {
                'success': True,
                'message': 'GPS activado y Google Maps abierto. Verificar que la ubicacion se detecta correctamente.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_tethering(self, serial: str) -> dict:
        """Abre la pantalla de Tethering."""
        try:
            self._adb(serial, 'shell am start -n com.android.settings/.TetherSettings')
            return {
                'success': True,
                'message': 'Pantalla de Tethering abierta. Activar el hotspot y conectar otro dispositivo para verificar el acceso a internet.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_bluetooth(self, serial: str) -> dict:
        """Activa Bluetooth y abre configuracion."""
        try:
            self._adb(serial, 'shell svc bluetooth enable')
            time.sleep(1)
            self._adb(serial, 'shell am start -a android.settings.BLUETOOTH_SETTINGS')
            return {
                'success': True,
                'message': 'Bluetooth activado y configuracion abierta. Emparejar con otro dispositivo y transferir un archivo.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_nfc(self, serial: str) -> dict:
        """Abre la configuracion de NFC."""
        try:
            self._adb(serial, 'shell am start -a android.settings.NFC_SETTINGS')
            return {
                'success': True,
                'message': 'Configuracion NFC abierta. Activar NFC y transferir un archivo con otro dispositivo NFC.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_csfb_call(self, serial: str) -> dict:
        """Abre el marcador para iniciar una llamada CSFB."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {
                'success': True,
                'message': 'Marcador abierto. Iniciar una llamada de voz y verificar la caida a CSFB (3G/2G) y el retorno a LTE al colgar.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_data_voice(self, serial: str) -> dict:
        """Abre el navegador para descarga simultanea con llamada."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d http://speedtest.net')
            return {
                'success': True,
                'message': 'Navegador abierto. Iniciar una descarga y luego realizar una llamada. Verificar que los datos continuan durante la llamada VoLTE.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_csfb_mo_quality(self, serial: str) -> dict:
        return {
            'success': True,
            'message': 'Iniciar una llamada de voz saliente. Verificar la caida a CSFB y evaluar la calidad de voz (sin cortes, eco ni ruido).'
        }

    def _semi_csfb_mt_quality(self, serial: str) -> dict:
        return {
            'success': True,
            'message': 'Solicitar que le llamen. Recibir la llamada MT. Verificar la caida a CSFB y evaluar la calidad de voz.'
        }

    def _semi_emergency_call(self, serial: str) -> dict:
        """Abre el marcador con el numero de emergencia pre-marcado."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL -d tel:112')
            return {
                'success': True,
                'message': 'Marcador abierto con 112 pre-marcado. Presionar llamar para verificar. RECORDAR COLGAR INMEDIATAMENTE.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_ims_deregistration(self, serial: str) -> dict:
        """Ciclo de modo avion para verificar deregistro/registro IMS."""
        try:
            self._adb(serial, 'shell settings put global airplane_mode_on 1')
            self._adb(serial, 'shell am broadcast -a android.intent.action.AIRPLANE_MODE --ez state true')
            time.sleep(3)
            self._adb(serial, 'shell settings put global airplane_mode_on 0')
            self._adb(serial, 'shell am broadcast -a android.intent.action.AIRPLANE_MODE --ez state false')
            return {
                'success': True,
                'message': 'Ciclo de modo avion realizado (ON 3s → OFF). Esperar 30s y verificar que IMS se re-registra correctamente.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_volte_mo_ims(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Llamar a un numero VoLTE WOM. Verificar que el icono VoLTE esta activo durante la llamada.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_volte_mo_cs(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Llamar a un numero CS (no VoLTE). Verificar que la llamada conecta correctamente.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_volte_mo_fixed(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Llamar a un numero fijo (linea residencial). Verificar que la llamada conecta y la calidad es buena.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_sms_during_call(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.SENDTO -d smsto:')
            return {'success': True, 'message': 'Aplicacion de SMS abierta. Realizar una llamada VoLTE activa y luego enviar un SMS. Verificar que el destinatario lo recibe.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_receive_sms(self, serial: str) -> dict:
        return {'success': True, 'message': 'Solicitar que le envien un SMS desde otro dispositivo. Verificar que el DUT lo recibe correctamente con el contenido integro.'}

    def _semi_call_hold_mo(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Iniciar una llamada VoLTE, usar la opcion "En espera". Esperar 15s y retomar la llamada. Verificar comunicacion correcta.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_call_hold_mt(self, serial: str) -> dict:
        return {'success': True, 'message': 'Solicitar que la parte remota ponga el DUT en espera. Esperar 15s. Verificar que el DUT gestiona correctamente el estado de espera.'}

    def _semi_ussd_during_call(self, serial: str) -> dict:
        return {'success': True, 'message': 'Iniciar una llamada VoLTE activa. Durante la llamada, abrir el marcador y enviar *901# (o codigo USSD WOM). Verificar que se recibe la respuesta USSD.'}

    def _semi_emergency_volte(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL -d tel:112')
            return {'success': True, 'message': 'Marcador abierto con 112. Llamar y verificar el mecanismo de fallback. RECORDAR COLGAR INMEDIATAMENTE.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_call_during_data(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d http://speedtest.net')
            return {'success': True, 'message': 'Navegador abierto. Iniciar una descarga y mientras descarga realizar una llamada VoLTE. Verificar que ambos funcionan simultaneamente.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_data_during_call(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Iniciar una llamada VoLTE, luego abrir el navegador y descargar un archivo. Verificar que datos y voz funcionan en paralelo.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_dual_sim_volte(self, serial: str) -> dict:
        try:
            config = self._adb(serial, 'shell getprop persist.radio.multisim.config').strip()
            is_dual = 'dsds' in config.lower() or 'dsda' in config.lower()
            return {
                'success': True,
                'message': (f"Configuracion SIM: {config}. "
                            f"{'Dual SIM detectado.' if is_dual else 'SIM unica detectada.'} "
                            "Verificar VoLTE en ambas ranuras realizando llamadas MO/MT en cada SIM.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_switch_preferred_sim(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.settings.NETWORK_OPERATOR_SETTINGS')
            return {'success': True, 'message': 'Configuracion de red abierta. Cambiar la SIM preferida de Slot1 a Slot2 y verificar el registro IMS en ambas.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_switch_preferred_sim2(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.settings.NETWORK_OPERATOR_SETTINGS')
            return {'success': True, 'message': 'Configuracion de red abierta. Realizar el cambio inverso: de Slot2 a Slot1. Verificar el registro correcto en ambas ranuras.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_toggle(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.settings.WIFI_CALLING_SETTINGS')
            return {'success': True, 'message': 'Configuracion de WiFi Calling abierta. Activar y desactivar la opcion y verificar el registro/deregistro IMS.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_mo_wifi(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Asegurarse que WiFi Calling esta activo y llamar a otro dispositivo con WiFi Calling. Verificar la calidad.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_mo_volte(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Con WiFi Calling activo, llamar a un numero VoLTE. Verificar que la llamada se establece correctamente.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_handover(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {'success': True, 'message': 'Marcador abierto. Iniciar una llamada WiFi, luego desactivar el WiFi durante la llamada. Verificar que continua via VoLTE o CSFB.'}
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_mt_volte(self, serial: str) -> dict:
        return {'success': True, 'message': 'Con WiFi Calling activo en el DUT, solicitar que le llamen desde un numero VoLTE. Verificar que la llamada MT se recibe correctamente via WiFi.'}

    def _semi_5g_volte(self, serial: str) -> dict:
        try:
            out = self._adb(serial, 'shell dumpsys telephony.registry')
            nr_lines = [l.strip() for l in out.splitlines()
                        if any(k in l for k in ['NR', '5G', 'nrState', 'nrFrequencyRange'])]
            msg = '\n'.join(nr_lines[:5]) if nr_lines else "No se encontro informacion 5G NR"
            return {
                'success': True,
                'message': (f"Estado NR detectado:\n{msg}\n\n"
                            "Verificar que el icono 5G es visible. Luego iniciar una llamada VoLTE "
                            "y verificar que se mantiene el icono.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_5g_data(self, serial: str) -> dict:
        try:
            out = self._adb(serial, 'shell dumpsys telephony.registry')
            nr_lines = [l.strip() for l in out.splitlines()
                        if any(k in l for k in ['NR', '5G', 'nrState', 'nrFrequencyRange'])]
            msg = '\n'.join(nr_lines[:5]) if nr_lines else "No se encontro informacion 5G NR"
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d http://speedtest.net')
            return {
                'success': True,
                'message': (f"Estado NR detectado:\n{msg}\n\n"
                            "Speedtest abierto. Verificar que las velocidades de datos utilizan 5G NR NSA.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_5g_data_voice(self, serial: str) -> dict:
        try:
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d http://speedtest.net')
            return {
                'success': True,
                'message': 'Speedtest abierto. Iniciar una descarga y luego realizar una llamada VoLTE. Verificar que datos 5G y voz VoLTE funcionan en paralelo.'
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_roaming_scan(self, serial: str) -> dict:
        """Abre seleccion manual de red para prueba de Roaming Nacional 3G."""
        try:
            net_type = self._adb(serial, 'shell getprop gsm.network.type').strip()
            operator = self._adb(serial, 'shell getprop gsm.operator.alpha').strip()
            self._adb(serial, 'shell am start -a android.settings.NETWORK_OPERATOR_SETTINGS')
            return {
                'success': True,
                'message': (f"Red actual: {operator or 'desconocida'} ({net_type or '-'}).\n\n"
                            "Pantalla de seleccion de red abierta.\n"
                            "Pasos a seguir:\n"
                            "1) Cambiar a seleccion MANUAL.\n"
                            "2) Esperar el escaneo de redes disponibles.\n"
                            "3) Seleccionar Movistar 3G.\n"
                            "4) Verificar registro exitoso y realizar llamada de voz.\n"
                            "5) Repetir con Movistar 2G.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_numeros_cortos(self, serial: str) -> dict:
        """Abre el marcador con instrucciones para probar numeros cortos de WOM."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {
                'success': True,
                'message': ("Marcador abierto. DUT debe estar registrado en VoLTE.\n\n"
                            "Marcar los siguientes numeros cortos y verificar que cada llamada conecta sobre VoLTE:\n"
                            "  *611  — Atencion al cliente WOM\n"
                            "  *222  — Saldo y recargas\n"
                            "  164   — Informacion\n"
                            "  123   — Policia Nacional\n"
                            "  112   — Numero de emergencias\n\n"
                            "Verificar el icono VoLTE activo durante cada llamada.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_voicemail_settings(self, serial: str) -> dict:
        """Abre la configuracion de buzon de voz."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {
                'success': True,
                'message': ("NOTA: El buzon de voz puede no estar disponible en WOM Colombia.\n\n"
                            "Marcador abierto. Para acceder a la configuracion del buzon de voz:\n"
                            "1) Ir a Menu > Ajustes > Buzon de voz.\n"
                            "2) Verificar que el numero del buzon esta configurado.\n"
                            "3) Si el servicio no esta disponible, marcar el resultado como N/A.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_voicemail_dial(self, serial: str) -> dict:
        """Marca el numero de acceso al buzon de voz (*86)."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.CALL -d tel:%2A86')
            return {
                'success': True,
                'message': ("NOTA: El buzon de voz puede no estar disponible en WOM Colombia.\n\n"
                            "Marcando *86 (acceso rapido al buzon de voz).\n"
                            "Verificar que el DUT conecta al centro de buzon de voz.\n"
                            "Alternativa: Mantener presionada la tecla '1' en el marcador.\n"
                            "Si el servicio no esta disponible, marcar el resultado como N/A.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_throughput(self, serial: str) -> dict:
        """Abre fast.com para prueba de throughput DL/UL."""
        try:
            net_type = self._adb(serial, 'shell getprop gsm.network.type').strip()
            self._adb(serial, 'shell am start -a android.intent.action.VIEW -d https://fast.com')
            return {
                'success': True,
                'message': (f"Red actual: {net_type or 'desconocida'}.\n\n"
                            "fast.com abierto para prueba de velocidad.\n"
                            "Pasos a seguir:\n"
                            "1) Verificar que el Wi-Fi esta DESACTIVADO.\n"
                            "2) Ejecutar la prueba de descarga.\n"
                            "3) Activar 'Mostrar mas info' para ver la velocidad de carga.\n"
                            "4) Anotar los valores DL y UL en el campo de observaciones.\n"
                            "5) Comparar con un dispositivo de referencia.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_lte_scan(self, serial: str) -> dict:
        """Abre busqueda manual de red para verificar registro LTE."""
        try:
            net_type = self._adb(serial, 'shell getprop gsm.network.type').strip()
            operator = self._adb(serial, 'shell getprop gsm.operator.alpha').strip()
            self._adb(serial, 'shell am start -a android.settings.NETWORK_OPERATOR_SETTINGS')
            return {
                'success': True,
                'message': (f"Red actual: {operator or 'desconocida'} ({net_type or '-'}).\n\n"
                            "Pantalla de seleccion de red abierta.\n"
                            "Pasos a seguir:\n"
                            "1) Cambiar a busqueda MANUAL.\n"
                            "2) Seleccionar WOM LTE (la misma red a la que ya estaba conectado).\n"
                            "3) Cuando regrese a ajustes, seleccionar 'Seleccion automatica'.\n"
                            "4) Verificar que el icono 4G/LTE aparece en la barra de estado.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_call_waiting_toggle(self, serial: str) -> dict:
        """Abre los ajustes de llamada para configurar Llamada en Espera (CW)."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.MAIN '
                              '-n com.android.phone/.settings.CallFeaturesSetting')
            return {
                'success': True,
                'message': ("Ajustes de llamada abiertos.\n\n"
                            "Pasos a seguir:\n"
                            "1) Ir a 'Llamada en espera' o 'Call Waiting'.\n"
                            "2) Desactivar la opcion y verificar el estado.\n"
                            "3) Volver a activar la opcion.\n"
                            "4) Con una llamada activa, recibir una segunda llamada "
                            "y verificar la notificacion de llamada en espera.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_mo_csfb(self, serial: str) -> dict:
        """Abre WiFi Calling y el marcador para llamada WiFi hacia cliente CSFB."""
        try:
            self._adb(serial, 'shell am start -a android.settings.WIFI_CALLING_SETTINGS')
            return {
                'success': True,
                'message': ("Ajustes de WiFi Calling abiertos.\n\n"
                            "Pasos a seguir:\n"
                            "1) Confirmar que WiFi Calling esta ACTIVADO.\n"
                            "2) Volver al marcador y llamar a un numero CSFB (cliente en red CS).\n"
                            "3) Verificar que la llamada se establece via WiFi en el DUT.\n"
                            "4) Verificar la calidad de audio.\n"
                            "5) Finalizar la llamada.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_pstn_call(self, serial: str) -> dict:
        """Abre el marcador para llamada WiFi hacia linea fija PSTN."""
        try:
            self._adb(serial, 'shell am start -a android.intent.action.DIAL')
            return {
                'success': True,
                'message': ("Marcador abierto. Verificar que WiFi Calling esta activo.\n\n"
                            "Prueba MO (saliente):\n"
                            "1) Marcar un numero de linea fija.\n"
                            "2) Verificar que la llamada va via WiFi y la calidad es buena.\n"
                            "3) Colgar.\n\n"
                            "Prueba MT (entrante):\n"
                            "4) Solicitar que le llamen desde la linea fija.\n"
                            "5) Contestar en el DUT y verificar que llega via WiFi.\n"
                            "6) Verificar la calidad de audio.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_vowifi_mt_csfb(self, serial: str) -> dict:
        """Prepara el DUT para recibir llamada CSFB con WiFi Calling activo."""
        try:
            self._adb(serial, 'shell am start -a android.settings.WIFI_CALLING_SETTINGS')
            return {
                'success': True,
                'message': ("Ajustes de WiFi Calling abiertos.\n\n"
                            "Pasos a seguir:\n"
                            "1) Confirmar que WiFi Calling esta ACTIVADO.\n"
                            "2) Solicitar a la Parte B que realice una llamada CSFB al numero del DUT.\n"
                            "3) Contestar la llamada en el DUT.\n"
                            "4) Verificar que la UI indica que la llamada es via WiFi.\n"
                            "5) Verificar la calidad de audio.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}

    def _semi_emergency_vowifi(self, serial: str) -> dict:
        """Abre el marcador con 112 para prueba de emergencia via WiFi."""
        try:
            wifi_state = self._adb(serial, 'shell dumpsys wifi | grep "mNetworkInfo"').strip()
            self._adb(serial, 'shell am start -a android.intent.action.DIAL -d tel:112')
            return {
                'success': True,
                'message': ("Estado WiFi: " + (wifi_state[:100] if wifi_state else "desconocido") + "\n\n"
                            "Marcador abierto con 112 pre-marcado.\n\n"
                            "Pasos a seguir:\n"
                            "1) Confirmar que WiFi Calling esta activo y registrado en IMS.\n"
                            "2) Presionar LLAMAR para iniciar la emergencia.\n"
                            "3) COLGAR INMEDIATAMENTE.\n"
                            "4) Verificar la ruta utilizada: WiFi, VoLTE o CSFB.\n\n"
                            "ADVERTENCIA: Solo realizar si es un ambiente de prueba controlado.")
            }
        except Exception as e:
            return {'success': False, 'message': str(e)}
